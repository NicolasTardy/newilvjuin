#!/usr/bin/env python3
"""Génère les ILV crédit BUT pour les produits >= 300€ du fichier dépliant.

Usage:
    python3 generate_ilv_depliant.py [fichier_depliant.xlsx] [--prix-min 300]

Résultat : dossier ILV_depliant_YYYYMMDD/ dans le même répertoire que ce script.
"""
import json
import os
import re
import sys
import warnings
from datetime import date, datetime, timedelta
from pathlib import Path

import fitz  # PyMuPDF
import openpyxl
import pypdf
from pypdf.generic import (
    ArrayObject,
    DecodedStreamObject,
    DictionaryObject,
    NameObject,
    TextStringObject,
)

warnings.filterwarnings("ignore", category=UserWarning)

# ── Chemins ────────────────────────────────────────────────────────────────
BASE_DIR = Path(__file__).resolve().parent
CHAMPS_JSON = BASE_DIR / "champs_all.json"
FONTS_DIR = BASE_DIR / "fonts"
# Archivo SemiCondensed = même police que celle utilisée dans le template
# Téléchargée depuis le repo officiel Omnibus-Type/Archivo (variable font)
ARCHIVO_BOLD_TTF = FONTS_DIR / "ArchivoSemiCondensed-Bold.ttf"

# ── Mappings métier (miroir exact du webapp) ────────────────────────────────
BP_DOMAINE_MAP = {
    "ELECTRICAL": "ELECTROMENAGER",
    "MEUBLE": "MEUBLE",
    "DECORATION": "DECORATION",
}

BP_FAMILLE_MAP = {
    "ELECTRICAL|LAVAGE|LAVE LINGE": "LAVE-LINGE",
    "ELECTRICAL|LAVAGE|LAVE VAISSELLE": "LAVE-VAISSELLE",
    "ELECTRICAL|LAVAGE|SECHE LINGE": "SECHE-LINGE",
    "ELECTRICAL|FROID|RÉFRIGÉRATEUR": "REFRIGERATEUR",
    "ELECTRICAL|FROID|REFRIGERATEUR": "REFRIGERATEUR",
    "ELECTRICAL|FROID|REF US (SIDE/SIDE)": "REFRIGERATEUR US",
    "ELECTRICAL|FROID|COMBINÉ": "COMBINE",
    "ELECTRICAL|FROID|COMBINE": "COMBINE",
    "ELECTRICAL|FROID|CONGÉLATEUR": "CONGELATEUR",
    "ELECTRICAL|FROID|CONGELATEUR": "CONGELATEUR",
    "ELECTRICAL|FROID|CAVE A VIN": "CAVE A VIN",
    "ELECTRICAL|CUISSON ENCASTRABLE|FOUR": "FOUR",
    "ELECTRICAL|CUISSON ENCASTRABLE|TABLE DE CUISSON": "TABLE DE CUISSON",
    "ELECTRICAL|CUISSON ENCASTRABLE|ASPIRATION": "HOTTE",
    "ELECTRICAL|PEM - ASPI - AIR - FMO|ASPIRATEUR": "ASPIRATEUR",
    "ELECTRICAL|PEM - ASPI - AIR - FMO|FOUR MICRO-ONDES": "FMO (Four Micro-Ondes)",
    "ELECTRICAL|PEM - ASPI - AIR - FMO|CLIMATISATION": "CLIM",
}

BP_RAYON_MAP = {
    "ELECTRICAL|CUISINIERE": "CUISINIERE",
    "ELECTRICAL|PEM - ASPI - AIR - FMO": "PEM (Petit Electromenager)",
    "ELECTRICAL|TV": "TV",
    "ELECTRICAL|SON": None,
    "MEUBLE|SIEGE": "SIEGE",
    "MEUBLE|CAC - LITS": "CAC (Canape/Convertible)",
    "MEUBLE|MEUBLES DE SEJOUR": "SEJOUR",
    "MEUBLE|PETIT MEUBLE SEJOUR": "SEJOUR",
    "MEUBLE|LITERIE": "LITERIE (Meuble)",
    "MEUBLE|CUISINE - BUFFET": "CUISINE (meuble)",
    "MEUBLE|RANGEMENT MODULAIRE": "RANGEMENT",
    "MEUBLE|BIBLIO - RANGEMENT": "RANGEMENT",
    "MEUBLE|ARMOIRE - DRESSING": "RANGEMENT",
    "MEUBLE|BUREAU": "SEJOUR",
    "MEUBLE|TABLE ET CHAISE": "SEJOUR",
    # ── Décoration ──
    "DECORATION|LUMINAIRE": "LUMINAIRE",
    "DECORATION|OBJET DE DECORATION": "DECO",
    "DECORATION|TAPIS": "DECO",
    "DECORATION|TEXTILE JOUR": "DECO",
    "DECORATION|TEXTILE NUIT": "DECO",
    "DECORATION|UTILITAIRE": "DECO",
}

RAYON_TO_UNIVERS_ILV = {
    "PEM (Petit Electromenager)": "PEM", "EXPRESSO": "PEM",
    "FMO (Four Micro-Ondes)": "PEM", "ROBOTS": "PEM", "ASPIRATEUR": "PEM",
    "LAVE-LINGE": "GEM/TV", "LAVE-VAISSELLE": "GEM/TV", "SECHE-LINGE": "GEM/TV",
    "REFRIGERATEUR": "GEM/TV", "REFRIGERATEUR US": "GEM/TV", "COMBINE": "GEM/TV",
    "CONGELATEUR": "GEM/TV", "CUISINIERE": "GEM/TV", "TABLE DE CUISSON": "GEM/TV",
    "FOUR": "GEM/TV", "HOTTE": "GEM/TV", "CAVE A VIN": "GEM/TV",
    "CLIM": "GEM/TV", "TV": "GEM/TV",
    "SEJOUR": "MEUBLE", "RANGEMENT": "MEUBLE", "CUISINE (meuble)": "MEUBLE",
    "LITERIE": "LITERIE", "LITERIE (Meuble)": "LITERIE",
    "SIEGE": "SIEGE", "CAC (Canape/Convertible)": "SIEGE",
    "LUMINAIRE": "DECO", "DECO": "DECO",
}

STRATEGIE_ILV = [
    {"min":   80, "max":  400,  "univers": ["PEM","GEM/TV","MEUBLE","LITERIE","SIEGE","DECO"], "credit": "5x_sf"},
    {"min":  400, "max":  900,  "univers": ["PEM","GEM/TV","MEUBLE","LITERIE","SIEGE","DECO"], "credit": "10x_ir"},
    {"min":  900, "max": 1500,  "univers": ["PEM","GEM/TV","MEUBLE","LITERIE","SIEGE","DECO"], "credit": "20x_ir"},
    {"min": 1500, "max": 4000,  "univers": ["PEM","GEM/TV","MEUBLE","LITERIE","SIEGE","DECO"], "credit": "36x_si"},
    {"min": 4000, "max": 99999, "univers": ["PEM","GEM/TV","MEUBLE","LITERIE","SIEGE","DECO"], "credit": "48x_si"},
]

CREDIT_TYPES = {
    "5x_sf":  {"label": "5× Sans Frais",          "duree": 5,  "famille": "gratuit", "variants": ["gar", "sans"]},
    "10x_ir": {"label": "10× Int. Remboursés",     "duree": 10, "famille": "ir",      "variants": ["gar", "gar_liv", "sans"]},
    "20x_ir": {"label": "20× Int. Remboursés",     "duree": 20, "famille": "ir",      "variants": ["gar", "gar_liv", "sans"]},
    "36x_si": {"label": "36× Services Inclus",     "duree": 36, "famille": "payant",  "variants": ["gar", "gar_liv", "sans"]},
    "48x_si": {"label": "48× Services Inclus",     "duree": 48, "famille": "payant",  "variants": ["gar", "gar_liv", "sans"]},
    "60x_si": {"label": "60× Services Inclus",     "duree": 60, "famille": "payant",  "variants": ["gar", "gar_liv", "sans"]},
}

CREDITKEY_TO_SLUG = {
    "5x_sf": "5x", "10x_ir": "10x", "20x_ir": "20x",
    "36x_si": "36x", "48x_si": "48x", "60x_si": "60x",
}

# page dans le PDF template : 1=gar, 2=gar_liv, 3=sans
VARIANT_TO_PAGE = {"gar": 1, "gar_liv": 2, "sans": 3}

COEFF_MENSUELS = {
    5:  {1: None,    2: None,    3: None},
    10: {1: 0.10806, 2: 0.10692, 3: 0.10386},
    20: {1: 0.05736, 2: 0.05674, 3: 0.05373},
    36: {1: 0.03423, 2: 0.03423, 3: 0.03149},
    48: {1: 0.02738, 2: 0.02738, 3: 0.02457},
    60: {1: 0.02332, 2: 0.02332, 3: 0.02044},
}

TAEG = {
    # 5x = "sans frais" → TAEG affiché = 0 %. Le 8,44 % (TAEG_FICTIF_5X)
    # n'apparaît QUE dans l'exemple des mentions légales.
    5:  {1: "0 %",     2: "0 %",     3: "0 %"},
    10: {1: "18,63 %", 2: "15,73 %", 3: "8,60 %"},
    20: {1: "17,40 %", 2: "15,73 %", 3: "8,60 %"},
    36: {1: "15,05 %", 2: "15,05 %", 3: "8,60 %"},
    48: {1: "15,05 %", 2: "15,05 %", 3: "8,60 %"},
    60: {1: "15,05 %", 2: "15,05 %", 3: "8,60 %"},
}

# ── Mentions légales — constantes ──────────────────────────────────────────

TAUX_DEBITEUR = {
    5:  {1: 8.13,  2: 8.13,  3: 8.13},
    10: {1: 17.21, 2: 14.70, 3: 8.28},
    20: {1: 16.15, 2: 14.70, 3: 8.28},
    36: {1: 14.10, 2: 14.10, 3: 8.28},
    48: {1: 14.10, 2: 14.10, 3: 8.28},
    60: {1: 14.10, 2: 14.10, 3: 8.28},
}

TAEG_FICTIF_5X     = 8.44  # Mis à jour 04/05/2026
TAUX_DEBITEUR_5X   = 8.13  # Mis à jour 04/05/2026

ASSURANCE_TAUX_MENSUEL = {
    20: 0.00196,
    36: 0.00128,
    48: 2.63 / 1500,
    60: 2.63 / 1500,
}
TAEA_ASSURANCE = {20: "4,83", 36: "2,99", 48: "3,96", 60: "3,82"}

DATE_CONDITIONS = "01/04/2026"

# Coordonnées (x0, y_haut, x1, y_bas) de la zone ML dans le PDF de sortie
# (système de coordonnées PyMuPDF : origine en haut-gauche, y vers le bas)
ML_ZONES = {
    "5x":  (14, 432, 406, 558),
    "10x": (14, 408, 406, 550),
    "20x": (14, 408, 406, 562),
    "36x": (14, 428, 406, 558),
    "48x": (14, 424, 406, 553),
    "60x": (14, 432, 406, 562),
}

# ── Textes des mentions légales (source : NEW PLV.docx) ─────────────────────

_ML_OFFRE = {
    "5x":  "Offre de crédit accessoire à une vente de 80€ à 3000€ sur une durée de 5 mois, "
           "pour un achat de 80€ à 3000€. Le coût du crédit est pris en charge par votre magasin. "
           "Taux Annuel Effectif Global fixe : 0%.",
    "10x": "Offre de crédit accessoire à une vente de 160€ à 25 000€ sur une durée de 10 mois, "
           "pour un achat de 160€ à 25 000€. Taux Annuel Effectif Global fixe compris entre 8,60% "
           "et 18,63% selon le montant du crédit.",
    "20x": "Offre de crédit accessoire à une vente de 320€ à 25 000€ sur une durée de 20 mois, "
           "pour un achat de 320€ à 25 000€. Taux Annuel Effectif Global fixe compris entre 8,60% "
           "et 17,40% selon le montant du crédit.",
    "36x": "Offre de crédit accessoire à une vente de 580€ à 25 000€ sur une durée de 36 mois, "
           "pour un achat de 580€ à 25 000€. Taux Annuel Effectif Global fixe compris entre 8,60% "
           "et 15,05 % selon le montant du crédit.",
    "48x": "Offre de crédit accessoire à une vente de 770€ à 25 000€ sur une durée de 48 mois, "
           "pour un achat de 770€ à 25 000€. Taux Annuel Effectif Global fixe compris entre 8,60% "
           "et 15,05 % selon le montant du crédit.",
    "60x": "Offre de crédit accessoire à une vente de 960€ à 25 000€ sur une durée de 60 mois, "
           "pour un achat de 960€ à 25 000€. Taux Annuel Effectif Global fixe compris entre 8,60% "
           "et 15,05 % selon le montant du crédit.",
}

_ML_ASSURANCE_TMPL = (
    "Le coût mensuel de l'assurance facultative Décès, Perte Totale et Irréversible d'Autonomie, "
    "Maladie-Accident souscrite auprès de Cardif Assurances Vie et Cardif Assurances Risques Divers "
    "est de {ass_m} et s'ajoute au montant de la mensualité indiqué ci-dessus. "
    "Le coût total de cette assurance facultative est de {ass_tot}. "
    "Le taux annuel effectif de cette assurance (TAEA) est de {taea}."
)

_ML_SOUS_RESERVE_A = (   # 5x, 36x, 48x, 60x
    "Sous réserve d'étude et d'acceptation de votre dossier par BNP Paribas Personal Finance. "
    "Cetelem est une marque de BNP Paribas Personal Finance. "
    "Établissement de crédit – S.A. au capital de 634 574 115€, "
    "Siège social : 1, boulevard Haussmann 75009 Paris. "
    "542 097 902 RCS Paris (www.cetelem.fr). "
    "N° ORIAS : 07 023 128 (www.orias.fr). Vous disposez d'un droit de rétractation."
)
_ML_SOUS_RESERVE_B = (   # 10x, 20x
    "Sous réserve d'étude et d'acceptation de votre dossier par BNP Paribas Personal Finance. "
    "Cetelem est une marque de l'établissement de crédit BNP Paribas Personal Finance, "
    "SA au capital de 634 574 115€ – "
    "Siège social : 1, boulevard Haussmann 75009 Paris – "
    "542 097 902 RCS Paris (www.cetelem.fr). "
    "N° ORIAS : 07 023 128 (www.orias.fr). Vous disposez d'un droit de rétractation."
)
_ML_BON_ACHAT = (
    "(1) Bon d'achat nominatif envoyé par courriel dans les 15 jours après confirmation de la commande "
    "(hors cuisine) et acceptation du financement, utilisable une seule fois, donnant droit à une remise "
    "immédiate égale à la valeur du bon indiquée, pour un montant minimum d'achat à la valeur du bon, "
    "valable 3 mois après sa date d'émission, dans tous les magasins BUT de France métropolitaine "
    "(hors but.fr et Marketplace), non cumulable avec tout autre bon d'achat ou chèque cadeau carte BUT, "
    "hors produit exclusifs Internet, prestations de service et pièces détachées. "
    "Il ne peut être cédé à titre onéreux, ni donner lieu à un échange ou un remboursement, même partiel. "
    "Tout retour de produit donnant lieu à un remboursement sera décompté du montant total d'achat pris en compte."
)
_ML_PUBLICITE = (
    "Publicité diffusée par BUT INTERNATIONAL, 722 041 860 RCS Meaux – "
    "1, avenue Spinoza 77184 Emerainville – N° ORIAS : 10 055 338, "
    "en qualité d'intermédiaire en opérations de banques immatriculé dans la catégorie mandataire "
    "exclusif de BNP Paribas Personal Finance. BUT INTERNATIONAL apporte son concours à la réalisation "
    "d'opérations de crédit à la consommation sans agir en qualité de Prêteur."
)


def _fmt_eur(v_str: str) -> str:
    """Convertit une valeur formatée (avec chr(0x80)=€ CP1252) en Unicode €."""
    return v_str.replace(chr(0x80), "€")


# Valeurs FIXES de l'exemple représentatif par crédit (mentions légales PLV BUT).
# Ce sont des exemples légaux figés, INDÉPENDANTS du produit affiché.
_ML_EXEMPLE = {
    "5x":  dict(vente="250€",   mens="50€",    total="250,00€",   taeg="8,44%",  tdeb="8,13%",  interets="5€"),
    "10x": dict(vente="500€",   mens="54,03€", total="540,30€",   taeg="18,63%", tdeb="17,21%", interets="40,30€"),
    "20x": dict(vente="1000€",  mens="57,36€", total="1 147,20€", taeg="17,40%", tdeb="16,15%", interets="147,20€"),
    "36x": dict(vente="1 000€", mens="34,23€", total="1 232,28€", taeg="15,05%", tdeb="14,10%", interets="232,28€"),
    "48x": dict(vente="1 500€", mens="41,07€", total="1 971,15€", taeg="15,05%", tdeb="14,10%", interets="471,15€"),
    "60x": dict(vente="1 500€", mens="34,98€", total="2 098,80€", taeg="15,05%", tdeb="14,10%", interets="598,80€"),
}

# Valeurs FIXES de l'assurance facultative par crédit (20x, 36x, 48x, 60x).
_ML_ASSURANCE_VALS = {
    "20x": dict(m="1,96€", tot="39,20€",  taea="4,83%"),
    "36x": dict(m="1,28€", tot="46,08€",  taea="2,99%"),
    "48x": dict(m="2,63€", tot="126,24€", taea="3,96%"),
    "60x": dict(m="2,63€", tot="157,80€", taea="3,82%"),
}


def _build_ml_text(slug: str, vals: dict = None) -> str:
    """Construit les mentions légales EXACTES par type de crédit (PLV BUT).

    Les valeurs de l'exemple représentatif et de l'assurance sont FIXES
    (figées par la réglementation), donc INDÉPENDANTES du produit affiché.
    """
    offre = _ML_OFFRE.get(slug, "")
    ex    = _ML_EXEMPLE.get(slug, {})
    date  = DATE_CONDITIONS

    if slug == "5x":
        exemple = (
            f"Exemple pour un achat et un crédit accessoire à une vente de {ex['vente']} sur 5 mois, "
            f"vous remboursez 5 mensualités de {ex['mens']}. "
            f"Montant total dû (par l'emprunteur) : {ex['total']}. "
            f"Le coût du crédit (TAEG fixe : {ex['taeg']}, taux débiteur fixe de {ex['tdeb']}, "
            f"intérêts : {ex['interets']}) est pris en charge par votre magasin. "
            f"Conditions au {date}"
        )
    else:
        duree = {"10x": 10, "20x": 20, "36x": 36, "48x": 48, "60x": 60}.get(slug, "")
        exemple = (
            f"Exemple pour un achat et un crédit accessoire à une vente de {ex['vente']} "
            f"sur {duree} mois au TAEG fixe de {ex['taeg']} (taux débiteur fixe de {ex['tdeb']}), "
            f"vous remboursez {duree} mensualités de {ex['mens']}, "
            f"intérêts : {ex['interets']}, "
            f"montant total dû (par l'emprunteur) : {ex['total']}. "
            f"Conditions au {date}."
        )

    # Assurance facultative (20x, 36x, 48x, 60x) — valeurs fixes
    assurance = ""
    av = _ML_ASSURANCE_VALS.get(slug)
    if av:
        assurance = _ML_ASSURANCE_TMPL.format(ass_m=av["m"], ass_tot=av["tot"], taea=av["taea"])

    sous_reserve = _ML_SOUS_RESERVE_B if slug in ("10x", "20x") else _ML_SOUS_RESERVE_A
    bon_achat    = _ML_BON_ACHAT if slug in ("10x", "20x") else ""

    parts = [offre, exemple]
    if assurance:
        parts.append(assurance)
    parts.append(sous_reserve)
    if bon_achat:
        parts.append(bon_achat)
    parts.append(_ML_PUBLICITE)

    return " ".join(p for p in parts if p)

# Simulateur tarifs services [univers, famille, prixMin, prixMax, gar1, gar2, gar3, liv]
SIM_RAW = [
["Electromenager","LAVE-LINGE",0,150,39.99,49.99,-1,39],["Electromenager","LAVE-LINGE",150,250,69.99,79.99,-1,39],["Electromenager","LAVE-LINGE",250,400,89.99,99.99,-1,39],["Electromenager","LAVE-LINGE",400,450,99.99,109.99,-1,39],["Electromenager","LAVE-LINGE",450,550,119.99,129.99,-1,39],["Electromenager","LAVE-LINGE",550,99999999,139.99,149.99,-1,39],
["Electromenager","REFRIGERATEUR",0,150,19.99,29.99,-1,39],["Electromenager","REFRIGERATEUR",150,250,39.99,49.99,-1,39],["Electromenager","REFRIGERATEUR",250,300,59.99,69.99,-1,39],["Electromenager","REFRIGERATEUR",300,400,69.99,79.99,-1,39],["Electromenager","REFRIGERATEUR",400,500,89.99,99.99,-1,39],["Electromenager","REFRIGERATEUR",500,600,109.99,119.99,-1,39],["Electromenager","REFRIGERATEUR",600,750,129.99,139.99,-1,39],["Electromenager","REFRIGERATEUR",750,99999999,139.99,149.99,-1,39],
["Electromenager","REFRIGERATEUR US",0,900,139.99,149.99,-1,79],["Electromenager","REFRIGERATEUR US",900,950,159.99,169.99,-1,79],["Electromenager","REFRIGERATEUR US",950,1000,189.99,199.99,-1,79],["Electromenager","REFRIGERATEUR US",1000,99999999,219.99,229.99,-1,79],
["Electromenager","SECHE-LINGE",0,250,69.99,79.99,-1,39],["Electromenager","SECHE-LINGE",250,350,79.99,89.99,-1,39],["Electromenager","SECHE-LINGE",350,450,109.99,119.99,-1,39],["Electromenager","SECHE-LINGE",450,500,119.99,129.99,-1,39],["Electromenager","SECHE-LINGE",500,99999999,129.99,139.99,-1,39],
["Electromenager","COMBINE",0,200,29.99,39.99,-1,39],["Electromenager","COMBINE",200,300,69.99,79.99,-1,39],["Electromenager","COMBINE",300,400,79.99,89.99,-1,39],["Electromenager","COMBINE",400,600,109.99,119.99,-1,39],["Electromenager","COMBINE",600,850,119.99,129.99,-1,39],["Electromenager","COMBINE",850,99999999,139.99,149.99,-1,39],
["Electromenager","CONGELATEUR",0,200,39.99,49.99,-1,39],["Electromenager","CONGELATEUR",200,350,59.99,69.99,-1,39],["Electromenager","CONGELATEUR",350,500,89.99,99.99,-1,39],["Electromenager","CONGELATEUR",500,900,119.99,129.99,-1,39],["Electromenager","CONGELATEUR",900,99999999,139.99,149.99,-1,39],
["Electromenager","LAVE-VAISSELLE",0,100,39.99,49.99,-1,39],["Electromenager","LAVE-VAISSELLE",100,150,59.99,69.99,-1,39],["Electromenager","LAVE-VAISSELLE",150,200,79.99,89.99,-1,39],["Electromenager","LAVE-VAISSELLE",200,300,89.99,99.99,-1,39],["Electromenager","LAVE-VAISSELLE",300,400,99.99,109.99,-1,39],["Electromenager","LAVE-VAISSELLE",400,450,109.99,119.99,-1,39],["Electromenager","LAVE-VAISSELLE",450,500,119.99,129.99,-1,39],["Electromenager","LAVE-VAISSELLE",500,700,129.99,139.99,-1,39],["Electromenager","LAVE-VAISSELLE",700,99999999,139.99,159.99,-1,39],
["Electromenager","TABLE DE CUISSON",0,100,19.99,29.99,-1,39],["Electromenager","TABLE DE CUISSON",100,150,29.99,39.99,-1,39],["Electromenager","TABLE DE CUISSON",150,200,34.99,44.99,-1,39],["Electromenager","TABLE DE CUISSON",200,250,39.99,49.99,-1,39],["Electromenager","TABLE DE CUISSON",250,300,49.99,59.99,-1,39],["Electromenager","TABLE DE CUISSON",300,400,59.99,69.99,-1,39],["Electromenager","TABLE DE CUISSON",400,500,69.99,79.99,-1,39],["Electromenager","TABLE DE CUISSON",500,650,99.99,109.99,-1,39],["Electromenager","TABLE DE CUISSON",650,99999999,109.99,129.99,-1,39],
["Electromenager","CUISINIERE",0,100,10.99,20.99,-1,39],["Electromenager","CUISINIERE",100,150,19.99,29.99,-1,39],["Electromenager","CUISINIERE",150,200,24.99,34.99,-1,39],["Electromenager","CUISINIERE",200,250,29.99,39.99,-1,39],["Electromenager","CUISINIERE",250,300,39.99,49.99,-1,39],["Electromenager","CUISINIERE",300,400,59.99,69.99,-1,39],["Electromenager","CUISINIERE",400,500,69.99,89.99,-1,39],["Electromenager","CUISINIERE",500,750,89.99,109.99,-1,39],["Electromenager","CUISINIERE",750,900,99.99,119.99,-1,39],["Electromenager","CUISINIERE",900,1000,119.99,139.99,-1,39],["Electromenager","CUISINIERE",1000,99999999,139.99,159.99,-1,39],
["Electromenager","FOUR",0,150,19.99,29.99,-1,39],["Electromenager","FOUR",150,250,39.99,49.99,-1,39],["Electromenager","FOUR",250,300,44.99,54.99,-1,39],["Electromenager","FOUR",300,400,59.99,69.99,-1,39],["Electromenager","FOUR",400,500,69.99,89.99,-1,39],["Electromenager","FOUR",500,650,79.99,99.99,-1,39],["Electromenager","FOUR",650,99999999,89.99,109.99,-1,39],
["Electromenager","HOTTE",0,100,14.99,24.99,-1,39],["Electromenager","HOTTE",100,150,19.99,29.99,-1,39],["Electromenager","HOTTE",150,200,24.99,34.99,-1,39],["Electromenager","HOTTE",200,300,39.99,49.99,-1,39],["Electromenager","HOTTE",300,350,49.99,69.99,-1,39],["Electromenager","HOTTE",350,99999999,59.99,79.99,-1,39],
["Electromenager","TV",0,100,0,19.99,-1,39],["Electromenager","TV",100,150,0,39.99,-1,39],["Electromenager","TV",150,200,0,49.99,-1,39],["Electromenager","TV",200,300,0,69.99,-1,39],["Electromenager","TV",300,400,0,89.99,-1,39],["Electromenager","TV",400,500,0,119.99,-1,39],["Electromenager","TV",500,600,0,129.99,-1,39],["Electromenager","TV",600,700,0,139.99,-1,39],["Electromenager","TV",700,800,0,169.99,-1,39],["Electromenager","TV",800,900,0,189.99,-1,39],["Electromenager","TV",900,1000,0,209.99,-1,39],["Electromenager","TV",1000,1500,0,289.99,-1,39],["Electromenager","TV",1500,2000,0,339.99,-1,39],["Electromenager","TV",2000,99999999,0,389.99,-1,39],
["Electromenager","FMO (Four Micro-Ondes)",0,50,0,9.99,-1,39],["Electromenager","FMO (Four Micro-Ondes)",50,100,0,14.99,-1,39],["Electromenager","FMO (Four Micro-Ondes)",100,150,0,29.99,-1,39],["Electromenager","FMO (Four Micro-Ondes)",150,200,0,39.99,-1,39],["Electromenager","FMO (Four Micro-Ondes)",200,300,0,49.99,-1,39],["Electromenager","FMO (Four Micro-Ondes)",300,99999999,0,59.99,-1,39],
["Electromenager","ROBOTS",0,200,0,29.99,-1,39],["Electromenager","ROBOTS",200,300,0,39.99,-1,39],["Electromenager","ROBOTS",300,500,0,54.99,-1,39],["Electromenager","ROBOTS",500,99999999,0,74.99,-1,39],
["Electromenager","EXPRESSO",0,100,0,19.99,-1,39],["Electromenager","EXPRESSO",100,150,0,24.99,-1,39],["Electromenager","EXPRESSO",150,200,0,29.99,-1,39],["Electromenager","EXPRESSO",200,400,0,39.99,-1,39],["Electromenager","EXPRESSO",400,99999999,0,59.99,-1,39],
["Electromenager","ASPIRATEUR",0,50,0,14.99,-1,39],["Electromenager","ASPIRATEUR",50,100,0,24.99,-1,39],["Electromenager","ASPIRATEUR",100,150,0,34.99,-1,39],["Electromenager","ASPIRATEUR",150,200,0,49.99,-1,39],["Electromenager","ASPIRATEUR",200,250,0,59.99,-1,39],["Electromenager","ASPIRATEUR",250,300,0,64.99,-1,39],["Electromenager","ASPIRATEUR",300,400,0,79.99,-1,39],["Electromenager","ASPIRATEUR",400,500,0,99.99,-1,39],["Electromenager","ASPIRATEUR",500,600,0,114.99,-1,39],["Electromenager","ASPIRATEUR",600,99999999,0,134.99,-1,39],
["Electromenager","CAVE A VIN",0,200,0,49.99,-1,39],["Electromenager","CAVE A VIN",200,350,0,59.99,-1,39],["Electromenager","CAVE A VIN",350,500,0,69.99,-1,39],["Electromenager","CAVE A VIN",500,99999999,0,79.99,-1,39],
["Electromenager","PEM (Petit Electromenager)",0,100,0,10.9,0,39],["Electromenager","PEM (Petit Electromenager)",100,99999999,0,19.9,0,39],
["Electromenager","CLIM",0,99999999,0,75.9,-1,39],
["Meubles","SIEGE",0,50,6.99,9.99,0,49],["Meubles","SIEGE",50,100,19.99,24.99,0,49],["Meubles","SIEGE",100,150,29.99,34.99,0,49],["Meubles","SIEGE",150,200,39.99,49.99,0,49],["Meubles","SIEGE",200,250,49.99,59.99,0,69],["Meubles","SIEGE",250,300,59.99,79.99,0,69],["Meubles","SIEGE",300,450,69.99,89.99,0,69],["Meubles","SIEGE",450,550,79.99,99.99,0,95],["Meubles","SIEGE",550,650,89.99,119.99,0,95],["Meubles","SIEGE",650,750,99.99,129.99,0,95],["Meubles","SIEGE",750,1000,109.99,149.99,0,105],["Meubles","SIEGE",1000,1500,139.99,189.99,209.99,125],["Meubles","SIEGE",1500,2000,179.99,229.99,249.99,139],["Meubles","SIEGE",2000,3000,0,0,299.99,139],["Meubles","SIEGE",3000,99999999,0,0,349.99,139],
["Meubles","LITERIE (Meuble)",0,50,6.99,19.99,0,40],["Meubles","LITERIE (Meuble)",50,100,9.99,24.99,0,40],["Meubles","LITERIE (Meuble)",100,200,19.99,34.99,0,40],["Meubles","LITERIE (Meuble)",200,250,29.99,39.99,0,40],["Meubles","LITERIE (Meuble)",250,300,34.99,49.99,0,40],["Meubles","LITERIE (Meuble)",300,350,39.99,59.99,0,40],["Meubles","LITERIE (Meuble)",350,400,44.99,69.99,0,40],["Meubles","LITERIE (Meuble)",400,450,49.99,79.99,89.99,40],["Meubles","LITERIE (Meuble)",450,500,59.99,89.99,99.99,40],["Meubles","LITERIE (Meuble)",500,550,69.99,99.99,109.99,40],["Meubles","LITERIE (Meuble)",550,600,79.99,109.99,119.99,40],["Meubles","LITERIE (Meuble)",600,650,89.99,119.99,129.99,40],["Meubles","LITERIE (Meuble)",650,700,99.99,129.99,139.99,40],["Meubles","LITERIE (Meuble)",700,750,109.99,139.99,149.99,40],["Meubles","LITERIE (Meuble)",750,800,119.99,149.99,159.99,40],["Meubles","LITERIE (Meuble)",800,850,129.99,159.99,169.99,40],["Meubles","LITERIE (Meuble)",850,900,149.99,169.99,189.99,40],["Meubles","LITERIE (Meuble)",900,1000,0,0,209.99,40],["Meubles","LITERIE (Meuble)",1000,1500,0,0,229.99,40],["Meubles","LITERIE (Meuble)",1500,2000,0,0,249.99,40],["Meubles","LITERIE (Meuble)",2000,3000,0,0,269.99,40],["Meubles","LITERIE (Meuble)",3000,99999999,0,0,289.99,40],
["Meubles","SEJOUR",0,50,9.99,14.99,0,49],["Meubles","SEJOUR",50,100,14.99,24.99,0,49],["Meubles","SEJOUR",100,150,24.99,39.99,0,49],["Meubles","SEJOUR",150,200,29.99,49.99,0,49],["Meubles","SEJOUR",200,250,39.99,59.99,0,69],["Meubles","SEJOUR",250,300,49.99,69.99,0,69],["Meubles","SEJOUR",300,500,69.99,79.99,0,69],["Meubles","SEJOUR",500,750,79.99,89.99,0,95],["Meubles","SEJOUR",750,1000,89.99,99.99,0,105],["Meubles","SEJOUR",1000,1200,109.99,119.99,0,125],["Meubles","SEJOUR",1200,1500,129.99,139.99,0,125],["Meubles","SEJOUR",1500,2000,149.99,159.99,0,139],["Meubles","SEJOUR",2000,3000,179.99,189.99,0,139],["Meubles","SEJOUR",3000,99999999,209.99,219.99,0,139],
["Meubles","CAC (Canape/Convertible)",0,50,9.99,14.99,0,49],["Meubles","CAC (Canape/Convertible)",50,100,14.99,19.99,0,49],["Meubles","CAC (Canape/Convertible)",100,150,24.99,34.99,0,49],["Meubles","CAC (Canape/Convertible)",150,200,34.99,49.99,0,49],["Meubles","CAC (Canape/Convertible)",200,250,39.99,59.99,0,69],["Meubles","CAC (Canape/Convertible)",250,300,49.99,69.99,0,69],["Meubles","CAC (Canape/Convertible)",300,1000,79.99,89.99,0,69],["Meubles","CAC (Canape/Convertible)",1000,99999999,99.99,109.99,0,125],
["Meubles","RANGEMENT",0,50,6.99,9.99,0,49],["Meubles","RANGEMENT",50,100,14.99,19.99,0,49],["Meubles","RANGEMENT",100,150,19.99,29.99,0,49],["Meubles","RANGEMENT",150,300,29.99,39.99,0,49],["Meubles","RANGEMENT",300,99999999,39.99,59.99,0,69],
["Meubles","CUISINE (meuble)",0,1000,129.99,0,0,49],["Meubles","CUISINE (meuble)",1000,1500,149.99,0,0,125],["Meubles","CUISINE (meuble)",1500,2000,199.99,0,0,139],["Meubles","CUISINE (meuble)",2000,2500,249.99,0,0,139],["Meubles","CUISINE (meuble)",2500,3000,299.99,0,0,139],["Meubles","CUISINE (meuble)",3000,4000,349.99,0,0,139],["Meubles","CUISINE (meuble)",4000,5000,399.99,0,0,139],["Meubles","CUISINE (meuble)",5000,7000,0,599.99,0,139],["Meubles","CUISINE (meuble)",7000,10000,0,749.99,0,139],["Meubles","CUISINE (meuble)",10000,99999999,0,999.99,0,139],
["Literie","LITERIE",0,50,6.99,19.99,0,40],["Literie","LITERIE",50,100,9.99,24.99,0,40],["Literie","LITERIE",100,200,19.99,34.99,0,40],["Literie","LITERIE",200,250,29.99,39.99,0,40],["Literie","LITERIE",250,300,34.99,49.99,0,40],["Literie","LITERIE",300,350,39.99,59.99,0,40],["Literie","LITERIE",350,400,44.99,69.99,0,40],["Literie","LITERIE",400,450,49.99,79.99,89.99,40],["Literie","LITERIE",450,500,59.99,89.99,99.99,40],["Literie","LITERIE",500,550,69.99,99.99,109.99,40],["Literie","LITERIE",550,600,79.99,109.99,119.99,40],["Literie","LITERIE",600,650,89.99,119.99,129.99,40],["Literie","LITERIE",650,700,99.99,129.99,139.99,40],["Literie","LITERIE",700,750,109.99,139.99,149.99,40],["Literie","LITERIE",750,800,119.99,149.99,159.99,40],["Literie","LITERIE",800,850,129.99,159.99,169.99,40],["Literie","LITERIE",850,900,149.99,169.99,189.99,40],["Literie","LITERIE",900,1000,0,0,209.99,40],["Literie","LITERIE",1000,1500,0,0,229.99,40],["Literie","LITERIE",1500,2000,0,0,249.99,40],["Literie","LITERIE",2000,3000,0,0,269.99,40],["Literie","LITERIE",3000,99999999,0,0,289.99,40],
]

SIM_DONNEES = [
    {"univers": r[0], "famille": r[1], "prixMin": r[2], "prixMax": r[3],
     "gar1": r[4], "gar2": r[5], "gar3": r[6], "liv": r[7]}
    for r in SIM_RAW
]


# ── Fonctions métier ────────────────────────────────────────────────────────

def lookup_warranty_all(rayon_ilv, prix):
    """Retourne {gar1, gar2, gar3, livraison} pour ce rayon et ce prix.

    Ignore l'univers ILV (GEM/TV, MEUBLE...) car SIM_DONNEES utilise
    des univers différents (Electromenager, Meubles...) — le fallback
    par famille seule suffit (même comportement que le webapp).
    """
    for d in SIM_DONNEES:
        if d["famille"] == rayon_ilv and d["prixMin"] <= prix <= d["prixMax"]:
            return {
                "gar1": d["gar1"] if d["gar1"] > 0 else 0,
                "gar2": d["gar2"] if d["gar2"] > 0 else 0,
                "gar3": d["gar3"] if d["gar3"] > 0 else 0,
                "livraison": d["liv"] if d["liv"] > 0 else 0,
            }
    return {"gar1": 0, "gar2": 0, "gar3": 0, "livraison": 0}


def best_garantie(war_all):
    """Choisit le niveau de garantie le plus élevé disponible (gar2 = 3★ en priorité).
    Retourne (niveau, prix).
    """
    if war_all["gar2"] > 0:
        return 2, war_all["gar2"]
    if war_all["gar1"] > 0:
        return 1, war_all["gar1"]
    return 0, 0


def determine_credit(prix, univers_ilv):
    if not univers_ilv:
        return None
    for rule in STRATEGIE_ILV:
        if rule["min"] <= prix < rule["max"] and univers_ilv in rule["univers"]:
            return rule["credit"]
    return None


def determiner_tranche(montant):
    # Bornes alignées sur le fichier officiel EASY PLV BUT (feuille « gamme ») :
    # tranche 1 ≤ 3000,99€ · tranche 2 ≤ 6000,99€ · tranche 3 jusqu'à 25 000€.
    if montant <= 3000.99:
        return 1
    if montant <= 6000.99:
        return 2
    return 3


def calculer_credit(duree, famille, montant_finance):
    tranche = determiner_tranche(montant_finance)
    taeg = TAEG[duree][tranche]
    if famille == "gratuit" or COEFF_MENSUELS[duree][tranche] is None:
        mensualite = round(montant_finance / duree, 2)
        return {"mensualite": mensualite, "taeg": taeg, "cout_credit": 0,
                "montant_total_du": montant_finance, "interets_rembourses": None}
    coeff = COEFF_MENSUELS[duree][tranche]
    mensualite = round(montant_finance * coeff, 2)
    montant_total_du = round(mensualite * duree, 2)
    cout_credit = round(montant_total_du - montant_finance, 2)
    interets_rembourses = cout_credit if famille == "ir" else None
    return {"mensualite": mensualite, "taeg": taeg, "cout_credit": cout_credit,
            "montant_total_du": montant_total_du, "interets_rembourses": interets_rembourses}


# ── Coordonnées exactes des champs PDF A3 (positions invariables) ────
# CONVERSION : les widgets PDF sont en système PDF (origine bas-gauche).
# Pour fitz (origine haut-gauche), on convertit : y_fitz = page_height - y_pdf
# Page A3 = 1190.55 pt de haut. Coordonnées ci-dessous DÉJÀ converties pour fitz.
# Format: [x0, y0, x1, y1] en coordonnées fitz (origine haut-gauche).
_A3_FIELD_COORDS = {
    # Champs communs à tous les variants
    # PDF widget rect [x0, y_bot_pdf, x1, y_top_pdf] → fitz [x0, page_h-y_top, x1, page_h-y_bot]
    "designation":      [53.0,  359.0,  416.0,  652.0],   # PDF [53, 539, 416, 831] → fitz y=359-652
    "mensualite":       [418.0, 494.0,  535.0,  580.0],   # PDF [418, 611, 535, 696] → fitz y=494-580
    "taeg":             [182.0, 595.0,  357.0,  616.0],   # PDF [182, 575, 357, 595] → fitz y=595-616
    "montant":          [238.0, 635.0,  489.0,  656.0],   # PDF [238, 535, 489, 555] → fitz y=635-656
    "mentions_legales": [96.0,  273.0,  481.0,  302.0],   # PDF [96, 889, 481, 917] → fitz y=273-302

    # Variant -1 (gar_liv) : 2 champs services
    "prix_livraison_gl": [262.0, 782.0, 332.0, 806.0],    # PDF [262, 384, 332, 408] → fitz y=782-806
    "prix_garantie_gl":  [719.0, 782.0, 805.0, 806.0],    # PDF [719, 385, 805, 408] → fitz y=782-806

    # Variant -2 (gar) : 1 champ service
    "prix_garantie_g":   [505.0, 741.0, 613.0, 769.0],    # PDF [505, 421, 613, 449] → fitz y=741-769

    # Badge IR pour 10x et 20x
    "badge_ir":          [622.0, 185.0, 729.0, 254.0],    # PDF [622, 936, 729, 1005] → fitz y=185-254
}


def fmt_montant(v):
    """Formate un montant en "1 234,56 €" (espaces insécables)."""
    s = f"{v:.2f}".replace(".", ",")
    parts = s.split(",")
    entier = parts[0]
    # Séparateur de milliers
    n = len(entier)
    groups = []
    while n > 3:
        groups.insert(0, entier[n-3:n])
        n -= 3
    groups.insert(0, entier[:n])
    return " ".join(groups) + "," + parts[1] + " €"


def fmt_prix_service(v):
    return f"{v:.2f}".replace(".", ",") + "€"


def make_filename(rayon_ilv, ean, credit_key, variant):
    famille = re.sub(r"[^a-zA-Z0-9àâäéèêëïîôùûüç()\-]", "_", rayon_ilv or "autre")
    ean_safe = re.sub(r"[^a-zA-Z0-9]", "_", ean or "sans-ean")
    return f"{famille}_{ean_safe}_{credit_key}_{variant}.pdf"


def build_field_values(product, calc, duree, famille, variant, page):
    """Retourne {role: value} → même logique que buildValuesForPage() du webapp."""
    mens = calc["mensualite"]
    entier = int(mens)
    cents = str(round((mens - entier) * 100)).zfill(2)
    taeg_val = (calc["taeg"] or "0 %").strip()
    cout_val = ("pris en charge par le magasin" if famille == "gratuit"
                else fmt_montant(calc["cout_credit"]))
    total_val = fmt_montant(calc["montant_total_du"])

    # Règle doublons page 1 vs pages 2/3
    is_page1 = (page == 1)
    line1 = taeg_val
    line2 = "" if is_page1 else cout_val
    line3 = cout_val if is_page1 else total_val
    line4 = ""
    line5 = total_val if is_page1 else ""
    line6 = ""

    is_ir = (famille == "ir")
    if is_ir and calc["interets_rembourses"] is not None:
        badge_taeg = fmt_montant(calc["interets_rembourses"]).rstrip(" €").strip()
    else:
        badge_taeg = (calc["taeg"] or "").strip()

    return {
        "designation":          product["designation"],
        "_doublon_designation": product["designation"],
        "ean_code":             ("EAN : " + product["ean"] if product["ean"] else ""),
        "entier_mensu":         str(entier),
        "centimes":             cents,
        "mensu_line_1":         line1,
        "mensu_line_2":         line2,
        "mensu_line_3":         line3,
        "mensu_line_4":         line4,
        "mensu_line_5":         line5,
        "mensu_line_6":         line6,
        "badge_garantie":       fmt_prix_service(product["garantie"]) if product["garantie"] > 0 else "",
        "badge_taeg":           badge_taeg,
        "service_gauche":       fmt_prix_service(product["livraison"]) if product["livraison"] > 0 else "",
        "service_droite":       fmt_prix_service(product["garantie"]) if product["garantie"] > 0 else "",
    }


# ── PDF generation ──────────────────────────────────────────────────────────

def _fmt_pdf_number(value):
    return f"{float(value):.6f}".rstrip("0").rstrip(".")


def flatten_widget_appearances(writer, page):
    """Dessine les apparences des champs dans le contenu puis supprime les widgets."""
    annots_ref = page.get("/Annots")
    if not annots_ref:
        return

    annots = annots_ref.get_object()
    if not annots:
        return

    resources = page.get("/Resources")
    if resources is None:
        resources = DictionaryObject()
        page[NameObject("/Resources")] = resources
    else:
        resources = resources.get_object()

    xobjects = resources.get("/XObject")
    if xobjects is None:
        xobjects = DictionaryObject()
        resources[NameObject("/XObject")] = xobjects
    else:
        xobjects = xobjects.get_object()

    draw_ops = []
    for idx, annot_ref in enumerate(annots):
        annot = annot_ref.get_object()
        if annot.get("/Subtype") != "/Widget":
            continue
        ap = annot.get("/AP")
        if not ap or "/N" not in ap:
            continue

        ap_ref = ap["/N"]
        ap_obj = ap_ref.get_object()
        rect = [float(v) for v in annot.get("/Rect", [])]
        bbox = [float(v) for v in ap_obj.get("/BBox", [])]
        if len(rect) != 4 or len(bbox) != 4:
            continue
        bw = bbox[2] - bbox[0]
        bh = bbox[3] - bbox[1]
        if bw == 0 or bh == 0:
            continue

        sx = (rect[2] - rect[0]) / bw
        sy = (rect[3] - rect[1]) / bh
        tx = rect[0] - bbox[0] * sx
        ty = rect[1] - bbox[1] * sy

        name = NameObject(f"/FmWidget{idx}")
        xobjects[name] = ap_ref
        draw_ops.append(
            "q "
            f"{_fmt_pdf_number(sx)} 0 0 {_fmt_pdf_number(sy)} "
            f"{_fmt_pdf_number(tx)} {_fmt_pdf_number(ty)} cm "
            f"{name} Do Q"
        )

    if not draw_ops:
        return

    stream = DecodedStreamObject()
    stream.set_data(("\n".join(draw_ops) + "\n").encode("utf-8"))
    stream_ref = writer._add_object(stream)

    contents = page.get("/Contents")
    if contents is None:
        page[NameObject("/Contents")] = stream_ref
    elif isinstance(contents, ArrayObject):
        contents.append(stream_ref)
    else:
        page[NameObject("/Contents")] = ArrayObject([contents, stream_ref])

    # Les textes sont maintenant dessinés dans la page : les widgets ne doivent
    # plus être interactifs ni surlignés par Preview.
    page.pop(NameObject("/Annots"), None)


ROLE_FONT_OVERRIDES = {
    # Le champ du template est trop haut visuellement une fois aplati.
    "centimes": ("/ArchivoSemiCondensed-ExtraBold", 27.0),
    # Les valeurs doivent rester lisibles mais moins envahissantes que le
    # montant principal.
    "mensu_line_1": ("/ArchivoSemiCondensed-ExtraBold", 8.5),
    "mensu_line_2": ("/ArchivoSemiCondensed-ExtraBold", 8.5),
    "mensu_line_3": ("/ArchivoSemiCondensed-ExtraBold", 8.5),
    "mensu_line_4": ("/ArchivoSemiCondensed-ExtraBold", 8.5),
    "mensu_line_5": ("/ArchivoSemiCondensed-ExtraBold", 8.5),
    "mensu_line_6": ("/ArchivoSemiCondensed-ExtraBold", 8.5),
}


def force_widget_font_weight(
    page,
    field_name_to_role,
    default_font_name="/ArchivoSemiCondensed-ExtraBold",
):
    """Force les widgets sur une fonte grasse avec quelques tailles par rôle."""
    annots_ref = page.get("/Annots")
    if not annots_ref:
        return
    for annot_ref in annots_ref.get_object():
        annot = annot_ref.get_object()
        if annot.get("/Subtype") != "/Widget":
            continue
        da = annot.get("/DA")
        if not da:
            continue
        field_name = str(annot.get("/T") or "")
        role = field_name_to_role.get(field_name, "")
        font_name, font_size = ROLE_FONT_OVERRIDES.get(
            role,
            (default_font_name, None),
        )

        def repl(match):
            size = font_size if font_size is not None else float(match.group(1))
            return f"{font_name} {_fmt_pdf_number(size)} Tf"

        annot[NameObject("/DA")] = TextStringObject(re.sub(
            r"/[A-Za-z0-9_.-]+\s+([0-9.]+)\s+Tf",
            repl,
            str(da),
        ))


def generate_ilv_pdf(slug, page_num, field_name_to_role, role_values,
                     pdf_template_path, output_path):
    """Remplit la page page_num du template PDF et sauvegarde en output_path."""
    reader = pypdf.PdfReader(str(pdf_template_path))
    writer = pypdf.PdfWriter()
    writer.clone_reader_document_root(reader)

    # Construit le dict champ_name → valeur pour cette page
    fill = {}
    for field_name, role in field_name_to_role.items():
        val = role_values.get(role, "")
        if val:
            fill[field_name] = val

    # Le rendu web utilise un poids très gras pour les valeurs injectées.
    # On force donc les champs PDF sur la fonte ExtraBold embarquée avant de
    # générer leurs apparences, puis on aplatit les champs.
    force_widget_font_weight(writer.pages[page_num - 1], field_name_to_role)

    # Met à jour les champs sur la page cible uniquement
    writer.update_page_form_field_values(
        writer.pages[page_num - 1], fill, auto_regenerate=True
    )

    # Extrait uniquement cette page en conservant la racine du PDF.
    # Un simple add_page() perd l'AcroForm /DR, donc les polices embarquées
    # des champs (Archivo SemiCondensed) et certains lecteurs retombent sur
    # Helvetica.
    keep_idx = page_num - 1
    for idx in reversed(range(len(writer.pages))):
        if idx != keep_idx:
            writer.remove_page(idx)

    flatten_widget_appearances(writer, writer.pages[0])
    writer.root_object.pop(NameObject("/AcroForm"), None)

    with open(str(output_path), "wb") as f:
        writer.write(f)


# ── Parsing du fichier dépliant ─────────────────────────────────────────────

def parse_depliant(xlsx_path, prix_min=300.0):
    wb = openpyxl.load_workbook(str(xlsx_path))
    ws = wb.active

    # Auto-détection ligne d'en-têtes (row 1 ou 2)
    headers_row1 = [str(c or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    if "LIB PDTS" in headers_row1:
        header_idx = 1
    else:
        headers_row1 = [str(c or "").strip() for c in next(ws.iter_rows(min_row=2, max_row=2, values_only=True))]
        header_idx = 2

    headers = headers_row1
    col = {h: i for i, h in enumerate(headers)}

    def idx(name):
        return col.get(name, -1)

    i_domaine = idx("DOMAINE")
    i_rayon   = idx("RAYON")
    i_famille = idx("FAMILLE")
    i_ean     = idx("EAN_13")
    i_lib     = idx("LIB PDTS")
    i_gamme   = idx("GAMME")
    i_debut   = idx("DEBUT")
    i_fin     = idx("FIN")
    i_pv_fdr  = idx("PV FDR")
    i_pv_pub  = idx("PV PUB")

    today = date.today()
    products = []

    for row in ws.iter_rows(min_row=header_idx + 1, values_only=True):
        designation = str(row[i_lib] or "").strip() if i_lib >= 0 else ""
        if not designation:
            continue

        domaine_bp = str(row[i_domaine] or "").strip() if i_domaine >= 0 else ""
        rayon_bp   = str(row[i_rayon]   or "").strip() if i_rayon   >= 0 else ""
        famille_bp = str(row[i_famille] or "").strip() if i_famille >= 0 else ""
        ean        = str(row[i_ean]     or "").strip() if i_ean     >= 0 else ""
        gamme      = str(row[i_gamme]   or "").strip() if i_gamme   >= 0 else ""

        pv_pub = float(row[i_pv_pub]) if i_pv_pub >= 0 and row[i_pv_pub] else 0.0
        pv_fdr = float(row[i_pv_fdr]) if i_pv_fdr >= 0 and row[i_pv_fdr] else 0.0

        raw_debut = row[i_debut] if i_debut >= 0 else None
        raw_fin   = row[i_fin]   if i_fin   >= 0 else None
        debut = raw_debut.date() if hasattr(raw_debut, "date") else (
            raw_debut if isinstance(raw_debut, date) else None)
        fin   = raw_fin.date()   if hasattr(raw_fin,   "date") else (
            raw_fin   if isinstance(raw_fin,   date) else None)

        d_grace = (debut - timedelta(days=1)) if debut else None
        in_window = d_grace and fin and d_grace <= today <= fin

        if in_window and pv_pub > 0:
            prix = pv_pub
        elif pv_fdr > 0:
            prix = pv_fdr
        elif pv_pub > 0:
            prix = pv_pub
        else:
            continue

        if prix < prix_min:
            continue

        # Filtre domaine
        domaine_info = BP_DOMAINE_MAP.get(domaine_bp)
        if domaine_info is None and domaine_bp in BP_DOMAINE_MAP:
            continue  # domaine mappé à None = exclu
        if domaine_bp not in BP_DOMAINE_MAP:
            continue  # domaine inconnu

        # Mapping rayon ILV
        fkey = f"{domaine_bp}|{rayon_bp}|{famille_bp}"
        rkey = f"{domaine_bp}|{rayon_bp}"
        rayon_ilv = BP_FAMILLE_MAP.get(fkey) or BP_RAYON_MAP.get(rkey)
        if rayon_ilv is None:
            continue

        univers_ilv = RAYON_TO_UNIVERS_ILV.get(rayon_ilv)
        if not univers_ilv:
            continue

        products.append({
            "designation": designation, "ean": ean, "gamme": gamme,
            "prix": prix, "rayon_ilv": rayon_ilv, "univers_ilv": univers_ilv,
        })

    return products


# ── Mentions légales dynamiques ─────────────────────────────────────────────

def fmt_ml(v: float, ndigits: int = 2) -> str:
    """Formate un montant pour les ML : séparateur de milliers espace, décimale virgule."""
    if ndigits == 0:
        s = str(round(v))
    else:
        s = f"{v:.{ndigits}f}".replace(".", ",")
    parts = s.split(",")
    entier = parts[0]
    n = len(entier)
    groups = []
    while n > 3:
        groups.insert(0, entier[n - 3:n])
        n -= 3
    groups.insert(0, entier[:n])
    result = " ".join(groups)
    if len(parts) > 1:
        result += "," + parts[1]
    return result


_CETELEM_A = (
    "Sous reserve d'etude et d'acceptation de votre dossier par BNP Paribas Personal Finance. "
    "Cetelem est une marque de BNP Paribas Personal Finance. "
    "Etablissement de credit - S.A. au capital de 634 574 115 EUR, "
    "Siege social : 1, boulevard Haussmann 75009 Paris. "
    "542 097 902 RCS Paris (www.cetelem.fr). "
    "N° ORIAS : 07 023 128 (www.orias.fr). "
    "Vous disposez d'un droit de retractation."
)
_CETELEM_B = (
    "Sous reserve d'etude et d'acceptation de votre dossier par BNP Paribas Personal Finance. "
    "Cetelem est une marque de l'etablissement de credit BNP Paribas Personal Finance, "
    "SA au capital de 634 574 115 EUR - "
    "Siege social : 1, boulevard Haussmann 75009 Paris - "
    "542 097 902 RCS Paris (www.cetelem.fr). "
    "N° ORIAS : 07 023 128 (www.orias.fr). "
    "Vous disposez d'un droit de retractation."
)
_BUT_PUB = (
    "Publicite diffusee par BUT INTERNATIONAL, 722 041 860 RCS Meaux - "
    "1, avenue Spinoza 77184 Emerainville - N° ORIAS : 10 055 338, "
    "en qualite d'intermediaire en operations de banques immatricule dans la categorie "
    "mandataire exclusif de BNP Paribas Personal Finance. BUT INTERNATIONAL apporte son "
    "concours a la realisation d'operations de credit a la consommation sans agir en "
    "qualite de Preteur."
)
_BON_ACHAT = (
    "(1) Bon d'achat nominatif envoye par courriel dans les 15 jours apres confirmation "
    "de la commande (hors cuisine) et acceptation du financement, utilisable une seule fois, "
    "donnant droit a une remise immediate egale a la valeur du bon indiquee, pour un montant "
    "minimum d'achat a la valeur du bon, valable 3 mois apres sa date d'emission, dans tous "
    "les magasins BUT de France metropolitaine (hors but.fr et Marketplace), non cumulable "
    "avec tout autre bon d'achat ou cheque cadeau carte BUT, hors produit exclusifs Internet, "
    "prestations de service et pieces detachees. Il ne peut etre cede a titre onereux, ni "
    "donner lieu a un echange ou un remboursement, meme partiel. Tout retour de produit "
    "donnant lieu a un remboursement sera decompte du montant total d'achat pris en compte."
)


def generer_ml_text(slug: str, duree: int, famille: str,
                    montant_finance: float, calc: dict) -> str:
    """Génère le texte des mentions légales dynamiques."""
    tranche   = determiner_tranche(montant_finance)
    taeg      = TAEG[duree][tranche].strip()
    taux_deb  = TAUX_DEBITEUR[duree][tranche]
    tdb_fmt   = f"{taux_deb:.2f}".replace(".", ",")
    mf_fmt    = fmt_ml(montant_finance)
    mens_fmt  = fmt_ml(calc["mensualite"])
    total_fmt = fmt_ml(calc["montant_total_du"])

    if slug == "5x":
        interets_fictifs = round(
            montant_finance * (TAUX_DEBITEUR_5X / 100 / 12) * ((duree + 1) / 2)
        )
        taeg_f_fmt = f"{TAEG_FICTIF_5X:.2f}".replace(".", ",")
        tdb_f_fmt  = f"{TAUX_DEBITEUR_5X:.2f}".replace(".", ",")
        return (
            "Offre de credit accessoire a une vente de 80EUR a 3000EUR sur une duree de 5 mois, "
            "pour un achat de 80EUR a 3000EUR. Le cout du credit est pris en charge par votre magasin. "
            "Taux Annuel Effectif Global fixe : 0 %. "
            f"Exemple pour un achat et un credit accessoire a une vente de {mf_fmt} EUR "
            f"sur 5 mois, vous remboursez 5 mensualites de {mens_fmt} EUR. "
            f"Montant total du (par l'emprunteur) : {total_fmt} EUR. "
            f"Le cout du credit (TAEG fixe : {taeg_f_fmt} %, "
            f"taux debiteur fixe de {tdb_f_fmt} %, "
            f"interets : {interets_fictifs} EUR) est pris en charge par votre magasin. "
            f"Conditions au {DATE_CONDITIONS}. "
            + _CETELEM_A + " " + _BUT_PUB
        )

    cout_fmt = fmt_ml(calc["cout_credit"])

    if slug == "10x":
        intro = (
            "Offre de credit accessoire a une vente de 160EUR a 25 000EUR sur une duree "
            "de 10 mois, pour un achat de 160EUR a 25 000EUR. Taux Annuel Effectif Global "
            "fixe compris entre 8,60 % et 18,63 % selon le montant du credit. "
        )
        exemple = (
            f"Exemple pour un achat et un credit accessoire a une vente de {mf_fmt} EUR "
            f"sur 10 mois au TAEG fixe de {taeg} (taux debiteur fixe de {tdb_fmt} %), "
            f"vous remboursez 10 mensualites de {mens_fmt} EUR, "
            f"interets : {cout_fmt} EUR, "
            f"montant total du (par l'emprunteur) : {total_fmt} EUR. "
            f"Conditions au {DATE_CONDITIONS}."
        )
        return intro + exemple + " " + _CETELEM_B + " " + _BON_ACHAT + " " + _BUT_PUB

    if slug == "20x":
        assurance_m  = round(montant_finance * ASSURANCE_TAUX_MENSUEL[duree], 2)
        cout_gar     = round(assurance_m * duree, 2)
        taea         = TAEA_ASSURANCE[duree]
        intro = (
            "Offre de credit accessoire a une vente de 320EUR a 25 000EUR sur une duree "
            "de 20 mois, pour un achat de 320EUR a 25 000EUR. Taux Annuel Effectif Global "
            "fixe compris entre 8,60 % et 17,40 % selon le montant du credit. "
        )
        exemple = (
            f"Exemple pour un achat et un credit accessoire a une vente de {mf_fmt} EUR "
            f"sur 20 mois au TAEG fixe de {taeg} (taux debiteur fixe de {tdb_fmt} %), "
            f"vous remboursez 20 mensualites de {mens_fmt} EUR, "
            f"interets : {cout_fmt} EUR, "
            f"montant total du (par l'emprunteur) : {total_fmt} EUR. "
            f"Conditions au {DATE_CONDITIONS}."
        )
        assurance_para = (
            f" Le cout mensuel de l'assurance facultative Deces, Perte Totale et Irreversible "
            f"d'Autonomie, Maladie-Accident souscrite aupres de Cardif Assurances Vie et "
            f"Cardif Assurances Risques Divers est de {fmt_ml(assurance_m)} EUR et s'ajoute "
            f"au montant de la mensualite indique ci-dessus. Le cout total de cette assurance "
            f"facultative est de {fmt_ml(cout_gar)} EUR. Le taux annuel effectif de cette "
            f"assurance (TAEA) est de {taea} %."
        )
        return intro + exemple + assurance_para + " " + _CETELEM_B + " " + _BON_ACHAT + " " + _BUT_PUB

    # 36x / 48x / 60x
    assurance_m = round(montant_finance * ASSURANCE_TAUX_MENSUEL[duree], 2)
    cout_gar    = round(assurance_m * duree, 2)
    taea        = TAEA_ASSURANCE[duree]
    taeg_max    = {"36x": "15,05", "48x": "15,05", "60x": "15,05"}[slug]
    montant_min = {"36x": "580", "48x": "770", "60x": "960"}[slug]
    intro = (
        f"Offre de credit accessoire a une vente de {montant_min}EUR a 25 000EUR "
        f"sur une duree de {duree} mois, pour un achat de {montant_min}EUR a 25 000EUR. "
        f"Taux Annuel Effectif Global fixe compris entre 8,60 % et {taeg_max} % "
        f"selon le montant du credit. "
    )
    exemple = (
        f"Exemple pour un achat et un credit accessoire a une vente de {mf_fmt} EUR "
        f"sur {duree} mois au TAEG fixe de {taeg} (taux debiteur fixe de {tdb_fmt} %), "
        f"vous remboursez {duree} mensualites de {mens_fmt} EUR, "
        f"interets : {cout_fmt} EUR, "
        f"montant total du (par l'emprunteur) : {total_fmt} EUR. "
        f"Conditions au {DATE_CONDITIONS}."
    )
    assurance_para = (
        f" Le cout mensuel de l'assurance facultative Deces, Perte Totale et Irreversible "
        f"d'Autonomie, Maladie-Accident souscrite aupres de Cardif Assurances Vie et "
        f"Cardif Assurances Risques Divers est de {fmt_ml(assurance_m)} EUR et s'ajoute "
        f"au montant de la mensualite indique ci-dessus. Le cout total de cette assurance "
        f"facultative est de {fmt_ml(cout_gar)} EUR. Le taux annuel effectif de cette "
        f"assurance (TAEA) est de {taea} %."
    )
    return intro + exemple + assurance_para + " " + _CETELEM_A + " " + _BUT_PUB


def ajouter_ml_overlay(output_path: str, slug: str, ml_text: str) -> None:
    """Écrase la zone ML statique et injecte le texte ML dynamique via PyMuPDF."""
    import tempfile

    if slug not in ML_ZONES:
        return

    # 1. Re-écrire le PDF proprement avec pypdf (nettoie la structure issue du clone)
    reader = pypdf.PdfReader(str(output_path))
    clean_writer = pypdf.PdfWriter()
    clean_writer.add_page(reader.pages[0])
    with open(str(output_path), "wb") as f:
        clean_writer.write(f)

    # 2. Appliquer l'overlay ML avec fitz
    x0, y0, x1, y1 = ML_ZONES[slug]
    rect = fitz.Rect(x0, y0, x1, y1)

    doc  = fitz.open(str(output_path))
    page = doc[0]

    # Rectangle blanc pour masquer le texte ML statique existant
    page.draw_rect(rect, color=(1, 1, 1), fill=(1, 1, 1))

    # Conversion Unicode → WinAnsi (CP1252) pour Helvetica built-in
    ml_win = ml_text.replace("€", chr(0x80))    # € → CP1252 0x80
    ml_win = ml_win.replace("–", chr(0x96)) # – (en-dash) → CP1252 0x96
    ml_win = ml_win.replace(" ", " ")       # espace fine insécable → espace normale

    # Injection du texte ML dynamique
    rc = page.insert_textbox(
        rect,
        ml_win,
        fontsize=5.4,
        fontname="helv",
        align=0,
        color=(0, 0, 0),
        encoding=fitz.TEXT_ENCODING_LATIN,
    )
    if rc < 0:
        print(f"      ⚠ ML overflow ({slug}) : {abs(rc):.0f} pt non affiché")

    # Sauvegarder vers un fichier temporaire puis remplacer l'original
    tmp_fd, tmp_path = tempfile.mkstemp(suffix=".pdf")
    os.close(tmp_fd)
    try:
        doc.save(tmp_path, incremental=False, garbage=3, deflate=True)
        doc.close()
        os.replace(tmp_path, str(output_path))
    except Exception:
        doc.close()
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)
        raise


# ── A3 Templates ────────────────────────────────────────────────────────────

A3_TEMPLATES_DIR = BASE_DIR / "PDF_MODIFIABLES_CREDIT"

# Chaque crédit a son sous-dossier (XY-A3) avec 3 fichiers variant (-1, -2, -3)
A3_TEMPLATE_FILES = {
    "5x":  "5X-A3/BDX_2610-DIVERS-ILV-CREDIT-5X-A3",
    "10x": "10X-A3/BDX_2610-DIVERS-ILV-CREDIT-10X-A3",
    "20x": "20X-A3/BDX_2610-DIVERS-ILV-CREDIT-20X-A3",
    "36x": "36X-A3/BDX_2610-DIVERS-ILV-CREDIT-36X-A3",
    "48x": "48X-A3/BDX_2610-DIVERS-ILV-CREDIT-48X-A3",
    "60x": "60X-A3/BDX_2610-DIVERS-ILV-CREDIT-60X-A3",
}

# Suffixe fichier par variante (chaque variant = fichier séparé)
A3_VARIANT_TO_SUFFIX = {"gar_liv": "-1", "gar": "-2", "sans": "-3"}

# Specs zone principale : (x_min, x_max, y_min, y_max, pat_substr, val_key, font_res, fontsize)
# Pour templates PDF_MODIFIABLES_CREDIT : cherche les LABELS et insère à côté
_A3_MAIN_SPECS = [
    # Mensualité: cherche "/ mois" (y≈466)
    (400, 800, 450, 500, "/ mois",    "mensualite_entier", "TT1", 95.0),
    # TAEG: cherche "TAEG fixe :" (y≈588)
    (50,  250, 580, 620, "TAEG fixe", "taeg_str",          "TT2", 22.0),
    # Montant total: cherche "Montant total dû :" (y≈628)
    (50,  300, 620, 665, "Montant",   "total_du_str",      "TT2", 22.0),
    # désignation : détectée par taille de police seule (sz≈26)
    (89,  452, 265, 310, None,        "designation",       "TT2", 26.2),
]

# Badge IR (10x, 20x) : ArchivoExtraCondensed-Black
_A3_IR_SPEC = (620, 735, 165, 290, "000", "ir_badge_str", "TT5", 78.0)

# Centimes — recherche "€" pour centimes (y≈187 ou y≈303)
_A3_CENTIMES_SPEC = (400, 550, 180, 320, "€", "centimes_str", "TT1", 95.0)

# Grand "€" statique — NOT USED for PDF_MODIFIABLES (€ est inséré avec centimes)
_A3_EURO_SPEC = (415, 510, 295, 315, "€", "TT1", 144.1)

# Service badges — page gar_liv (index 0) uniquement
_A3_SERVICE_SPECS = [
    (240, 310, 773, 816, "à 00€",  "livraison_fmt", "TT4", 26.6),
    (695, 775, 772, 816, "à000€",  "garantie_fmt",  "TT4", 26.6),
]

# Specs zone ML (inline, sz≈9) : (x_min, x_max, y_min, y_max, val_key)
_A3_ML_SPECS = {
    # Bornes X serrées : xmx juste avant le x0 du texte de contexte suivant
    # pour éviter de capturer et remplacer ce texte de contexte (doublons).
    "5x": [
        (256, 281, 1025, 1043, "ml_montant"),        # x0=261  — contexte suivant x0=282
        (451, 474, 1025, 1043, "ml_mensualite"),      # x0=456  — contexte suivant x0=475
        (612, 649, 1025, 1043, "ml_total_du"),        # x0=617  — contexte suivant x0=650
        (758, 784, 1025, 1043, "ml_taeg"),            # x0=763  — contexte suivant x0=786
        ( 87, 120, 1035, 1053, "ml_taux_deb"),        # x0=92   — contexte suivant x0=121
        (151, 181, 1035, 1053, "ml_interets"),        # x0=156  — contexte suivant x0=182
    ],
    "10x": [
        (256, 283,  985, 1001, "ml_montant"),         # x0=261  — contexte suivant x0=285
        (387, 422,  985, 1001, "ml_taeg"),            # x0=392  — contexte suivant x0=424
        (498, 528,  985, 1001, "ml_taux_deb"),        # x0=503  — contexte suivant x0=530
        (669, 703,  985, 1001, "ml_mensualite"),      # x0=674  — contexte suivant x0=705
        (736, 768,  985, 1001, "ml_interets"),        # x0=741  — contexte suivant x0=770
        (131, 172,  995, 1011, "ml_total_du"),        # x0=136  — contexte suivant x0=174
    ],
    "20x": [
        (253, 283,  965,  981, "ml_montant"),         # x0=258  — contexte suivant x0=285
        (384, 418,  965,  981, "ml_taeg"),            # x0=389  — contexte suivant x0=419
        (492, 528,  965,  981, "ml_taux_deb"),        # x0=497  — contexte suivant x0=530
        (667, 703,  965,  981, "ml_mensualite"),      # x0=672  — contexte suivant x0=705
        (737, 768,  965,  981, "ml_interets"),        # x0=742  — contexte suivant x0=770
        (133, 174,  975,  991, "ml_total_du"),        # x0=138  — contexte suivant x0=177
        (736, 767,  986, 1001, "ml_assurance_m"),     # x0=741  — contexte suivant x0=768
        (382, 415,  996, 1012, "ml_assurance_total"), # x0=387  — contexte suivant x0=417
        (621, 656,  996, 1012, "ml_taea"),            # x0=626  — contexte suivant x0=657
    ],
    # 36x, 48x, 60x : positions identiques — bornes X serrées pour ne pas
    # capturer le texte de contexte adjacent (ex: "sur X mois au TAEG fixe de")
    "36x": [
        (255, 297, 1005, 1025, "ml_montant"),        # placeholder x0=260, contexte suivant à x0=298
        (399, 434, 1005, 1025, "ml_taeg"),            # placeholder x0=404, contexte suivant à x0=439
        (513, 548, 1005, 1025, "ml_taux_deb"),        # placeholder x0=518, contexte suivant à x0=552
        (691, 728, 1005, 1025, "ml_mensualite"),      # placeholder x0=696, contexte suivant à x0=729
        (761, 803, 1005, 1025, "ml_interets"),        # placeholder x0=766, contexte suivant à x0=804
        (168, 212, 1016, 1040, "ml_total_du"),        # placeholder x0=173, contexte suivant à x0=214
        (736, 767, 1025, 1050, "ml_assurance_m"),     # placeholder x0=741, contexte suivant à x0=768
        (381, 417, 1036, 1062, "ml_assurance_total"), # placeholder x0=386, contexte suivant à x0=418
        (620, 656, 1036, 1062, "ml_taea"),            # placeholder x0=625, contexte suivant à x0=657
    ],
    "48x": [
        (255, 297, 1005, 1025, "ml_montant"),
        (399, 434, 1005, 1025, "ml_taeg"),
        (513, 548, 1005, 1025, "ml_taux_deb"),
        (691, 728, 1005, 1025, "ml_mensualite"),
        (761, 803, 1005, 1025, "ml_interets"),
        (168, 212, 1016, 1040, "ml_total_du"),
        (736, 767, 1025, 1050, "ml_assurance_m"),
        (381, 417, 1036, 1062, "ml_assurance_total"),
        (620, 656, 1036, 1062, "ml_taea"),
    ],
    "60x": [
        (255, 297, 1005, 1025, "ml_montant"),
        (399, 434, 1005, 1025, "ml_taeg"),
        (513, 548, 1005, 1025, "ml_taux_deb"),
        (691, 728, 1005, 1025, "ml_mensualite"),
        (761, 803, 1005, 1025, "ml_interets"),
        (168, 212, 1016, 1040, "ml_total_du"),
        (736, 767, 1025, 1050, "ml_assurance_m"),
        (381, 417, 1036, 1062, "ml_assurance_total"),
        (620, 656, 1036, 1062, "ml_taea"),
    ],
}


def _a3_compute_values(slug, variant, designation, calc, montant_finance,
                       garantie_prix, livraison_prix, duree, famille):
    """Calcule toutes les valeurs à substituer dans un PDF A3."""

    def feur(v, nd=2):
        return fmt_ml(v, nd) + chr(0x80)  # chr(0x80) = € en WinAnsiEncoding/CP1252

    def fpct(v):
        return f"{v:.2f}%".replace(".", ",")

    tranche = determiner_tranche(montant_finance)
    taeg    = TAEG[duree][tranche].replace(" %", "%")
    tdb     = TAUX_DEBITEUR[duree][tranche]

    mens         = calc["mensualite"]
    mens_entier  = int(mens)
    mens_cents   = f"{round((mens - mens_entier) * 100):02d}"
    total_str    = fmt_ml(calc["montant_total_du"]) + " " + chr(0x80)  # espaces normaux + € CP1252

    vals = {
        "mensualite_entier": str(mens_entier),
        "centimes_str":      mens_cents,
        "taeg_str":          taeg,
        "total_du_str":      total_str,
        "designation":       designation,
        "ir_badge_str":      None,
        "livraison_fmt":     "",
        "garantie_fmt":      "",
        # Zones dynamiques STC (template vectorisé)
        "desig":             designation,
        "liv_prix":          "",
        "gar_prix":          "",
    }

    if garantie_prix > 0:
        if variant == "gar_liv":
            # Pattern dans _A3_SERVICE_SPECS: "à000€" (sans espace pour 3 chiffres)
            vals["garantie_fmt"] = "à" + str(int(garantie_prix)).zfill(3) + chr(0x80)
            vals["gar_prix"] = str(int(garantie_prix)) + chr(0x80)
        elif variant == "gar":
            vals["garantie_fmt"] = str(int(garantie_prix)) + chr(0x80)
            vals["gar_prix"] = str(int(garantie_prix)) + chr(0x80)

    # Livraison : pour variante "gar" ET "gar_liv"
    if livraison_prix > 0 and variant in ("gar", "gar_liv"):
        # Pattern dans _A3_SERVICE_SPECS: "à 00€" (avec espace pour 2 chiffres)
        vals["livraison_fmt"] = "à " + str(int(livraison_prix)).zfill(2) + chr(0x80)
        vals["liv_prix"] = str(int(livraison_prix)) + chr(0x80)

    if slug in ("10x", "20x") and calc.get("interets_rembourses") is not None:
        vals["ir_badge_str"] = str(int(round(calc["interets_rembourses"])))

    # Zone ML
    if slug == "5x":
        m5    = int(round(montant_finance / duree))
        i_fic = round(montant_finance * (TAUX_DEBITEUR_5X / 100 / 12) * ((duree + 1) / 2))
        vals.update({
            "ml_montant":    feur(montant_finance, 0),
            "ml_mensualite": str(m5) + chr(0x80),
            "ml_total_du":   feur(calc["montant_total_du"]),
            "ml_taeg":       f"{TAEG_FICTIF_5X:.2f}%".replace(".", ","),
            "ml_taux_deb":   fpct(TAUX_DEBITEUR_5X),
            "ml_interets":   feur(i_fic),
        })
    else:
        vals.update({
            "ml_montant":    feur(montant_finance, 0),
            "ml_taeg":       taeg,
            "ml_taux_deb":   fpct(tdb),
            "ml_mensualite": feur(calc["mensualite"]),
            "ml_interets":   feur(calc["cout_credit"]),
            "ml_total_du":   feur(calc["montant_total_du"]),
        })
        if slug in ("20x", "36x", "48x", "60x"):
            ass_m   = round(montant_finance * ASSURANCE_TAUX_MENSUEL[duree], 2)
            ass_tot = round(ass_m * duree, 2)
            taea_v  = float(TAEA_ASSURANCE[duree].replace(",", "."))
            vals.update({
                "ml_assurance_m":     feur(ass_m),
                "ml_assurance_total": feur(ass_tot),
                "ml_taea":            fpct(taea_v),
            })

    return vals


def _identify_a3_field_roles(reader_page):
    """Identifie le rôle de chaque champ widget par sa position (Rect).

    Les noms des champs varient entre templates ("Champ de texte 17" ou
    "Champ de texte 2") mais leur position est invariable. Cette fonction
    retourne {field_name: role} basé sur les coordonnées (PDF, y du bas).
    """
    field_to_role = {}
    annots_ref = reader_page.get("/Annots")
    if not annots_ref:
        return field_to_role
    annots = annots_ref.get_object()

    for annot_ref in annots:
        annot = annot_ref.get_object()
        if annot.get("/Subtype") != "/Widget":
            continue
        name = str(annot.get("/T", ""))
        rect = annot.get("/Rect")
        if not name or not rect:
            continue
        try:
            x0, y0, x1, y1 = [float(v) for v in rect]
        except (TypeError, ValueError):
            continue
        w, h = x1 - x0, y1 - y0

        # Coordonnées en système PDF (origine bas-gauche).
        # Identification par position : zones invariables entre tous les templates.
        # Tolérance ±5px sur chaque coord pour absorber les petites variations.
        cx, cy = (x0 + x1) / 2, (y0 + y1) / 2

        # Désignation : BULLE DE DIALOGUE en haut (PDF coords 96-481 × 889-917)
        # ATTENTION: en système PDF, y haut = grand y. Donc ce champ EST en haut.
        # centre ~288,903
        if 270 < cx < 310 and 895 < cy < 910 and w > 350:
            field_to_role[name] = "designation"
        # Mensualité : grande zone milieu-droit (419-535 × 611-696) → centre ~477,654
        elif 460 < cx < 495 and 645 < cy < 665 and h > 80:
            field_to_role[name] = "mensualite"
        # TAEG : (182-357 × 575-595) → centre ~270,585
        elif 250 < cx < 290 and 580 < cy < 590 and 15 < h < 25:
            field_to_role[name] = "taeg"
        # Montant total dû : (238-489 × 535-555) → centre ~363,545
        elif 350 < cx < 380 and 540 < cy < 550 and 15 < h < 25:
            field_to_role[name] = "montant"
        # Mentions légales : grande zone gauche pour texte long
        # (53-416 × 539-831 PDF) → centre ~235,685
        elif 200 < cx < 260 and 680 < cy < 690 and w > 350:
            field_to_role[name] = "mentions_legales"
        # Prix livraison (gar_liv) : (262-332 × 384-408) → centre ~297,396
        elif 280 < cx < 310 and 390 < cy < 405:
            field_to_role[name] = "prix_livraison"
        # Prix garantie (gar_liv) : (719-805 × 385-408) → centre ~762,396
        elif 750 < cx < 780 and 390 < cy < 405:
            field_to_role[name] = "prix_garantie"
        # Prix garantie (gar seule) : (505-613 × 421-449) → centre ~559,435
        elif 540 < cx < 580 and 430 < cy < 445:
            field_to_role[name] = "prix_garantie"
        # Badge IR (10x, 20x) : (622-731 × 936-1005) → centre ~676,970
        elif 660 < cx < 700 and 960 < cy < 990:
            field_to_role[name] = "badge_ir"

    return field_to_role


# ── Zones A3 (coords fitz, origine haut-gauche) — invariables entre crédits ──
# Issues de l'inspection des widgets PDF, converties (y_fitz = 1190.55 - y_pdf).
# Cf. maquette annotée fournie par le client.
_A3_ZONES = {
    "designation":     (90,  264, 487, 306),   # bulle de dialogue — auto-size largeur (agrandie)
    "ean":             (78,  312, 360, 340),   # code EAN, juste sous la désignation
    "prix_produit":    (78,  342, 360, 365),   # prix produit hors services, sous le code EAN
    "mensu_int":       (50,  355, 414, 582),   # entier mensualité — auto-size (gros)
    "centimes":        (419, 492, 538, 583),   # centimes — taille fixe
    "prix_livraison":  (262, 783, 360, 807),   # prix livraison (gar_liv)
    "prix_garantie_gl":(719, 782, 815, 806),   # prix garantie (gar_liv)
    "prix_garantie_g": (505, 741, 625, 770),   # prix garantie (gar seule)
    "badge_ir":        (610, 185, 722, 240),   # badge intérêts remboursés (10x,20x) — intérieur du cercle
}

# Baselines fixes pour les valeurs alignées sur un label statique (size, x, baseline_y)
_A3_FIXED = {
    "taeg":           dict(x=184, baseline=614, size=22, max_w=170),   # à droite de "TAEG fixe :"
    "montant":        dict(x=240, baseline=654, size=22, max_w=245),   # à droite de "Montant total dû :"
    "liv":            dict(x=262, baseline=805, size=27, max_w=120),   # après "Livraison à" (baseline label ≈805)
    "gar_gl":         dict(x=719, baseline=804, size=27, max_w=120),   # après "Garantie ... à" (gar_liv)
    "gar_g":          dict(x=507, baseline=765, size=28, max_w=115),   # après "Garantie ... à" (gar)
}

# Couleurs
_COL_WHITE = (1, 1, 1)
_COL_RED   = (0.886, 0.0, 0.039)   # rouge BUT (#E2000A)
_COL_DARK  = (0.12, 0.12, 0.12)


def _a3_font():
    """Charge la police Archivo SemiCondensed Bold (même que le template)."""
    if ARCHIVO_BOLD_TTF.exists():
        return fitz.Font(fontfile=str(ARCHIVO_BOLD_TTF))
    return fitz.Font("helv")


def _draw_fitted(page, font, text, rect, color, max_size,
                 align="center", min_size=5, fill_h=0.92, fill_w=0.95):
    """Dessine `text` dans `rect`, auto-dimensionné pour tenir (centré V)."""
    if not text:
        return
    x0, y0, x1, y1 = rect
    bw, bh = x1 - x0, y1 - y0
    asc, desc = font.ascender, font.descender   # asc>0, desc<0 (par em)
    span = asc - desc

    size = max_size
    tw = font.text_length(text, size)
    if tw > bw * fill_w:
        size = size * (bw * fill_w) / tw
    if span * size > bh * fill_h:
        size = (bh * fill_h) / span
    size = max(size, min_size)

    tw = font.text_length(text, size)
    if align == "center":
        tx = x0 + (bw - tw) / 2
    elif align == "right":
        tx = x1 - tw
    else:
        tx = x0
    ty = y0 + (bh - span * size) / 2 + asc * size

    w = fitz.TextWriter(page.rect, color=color)
    w.append(fitz.Point(tx, ty), text, font=font, fontsize=size)
    w.write_text(page)


def _draw_at(page, font, text, x, baseline_y, size, color, max_w=None):
    """Dessine `text` à gauche depuis (x, baseline_y), rétréci si > max_w."""
    if not text:
        return
    if max_w:
        tw = font.text_length(text, size)
        if tw > max_w:
            size = size * max_w / tw
    w = fitz.TextWriter(page.rect, color=color)
    w.append(fitz.Point(x, baseline_y), text, font=font, fontsize=size)
    w.write_text(page)


def generate_a3_pdf(slug, variant, designation, calc, montant_finance,
                    garantie_prix, livraison_prix, duree, famille,
                    template_path, output_path, ean="", prix=0.0):
    """Génère un PDF A3 ILV — rendu fitz direct aux zones exactes des widgets.

    Layout (cf. maquette annotée client) :
      - Désignation  : bulle de dialogue (auto-size sur la longueur)
      - Mensualité   : ENTIER dans la grande boîte gauche (auto-size, gros)
                       + CENTIMES dans la petite boîte (taille fixe)
                       → le « € » et « / mois » sont STATIQUES (non touchés)
      - TAEG / Montant total dû : taille fixe, à droite de leur label
      - Prix livraison / garantie : taille fixe (selon variant), rouge
      - Mentions légales : grande boîte du bas (auto-size pour remplir)

    Variants : -1 gar_liv (liv+gar) · -2 gar (gar) · -3 sans.
    """
    vals = _a3_compute_values(slug, variant, designation, calc, montant_finance,
                               garantie_prix, livraison_prix, duree, famille)
    ml_text = _build_ml_text(slug, vals)

    mensu_entier = vals.get("mensualite_entier", "")
    centimes     = vals.get("centimes_str", "00")
    taeg         = vals.get("taeg_str", "")
    total_du     = _fmt_eur(vals.get("total_du_str", ""))
    desig        = vals.get("desig", designation or "")
    liv_prix     = _fmt_eur(vals.get("liv_prix", ""))
    gar_prix     = _fmt_eur(vals.get("gar_prix", ""))
    interets     = calc.get("interets_rembourses")

    # Ouvrir le template (1 page) avec fitz
    doc = fitz.open(str(template_path))
    page = doc[0]

    # Supprimer les widgets (placeholders vides) pour un rendu propre
    try:
        for wdg in list(page.widgets() or []):
            page.delete_widget(wdg)
    except Exception:
        pass

    font = _a3_font()
    Z = _A3_ZONES
    F = _A3_FIXED

    # Désignation — bulle claire → texte foncé, auto-size sur la longueur (agrandie)
    _draw_fitted(page, font, desig, Z["designation"], _COL_DARK,
                 max_size=34, align="center", fill_h=0.98)

    # Code EAN — sous la désignation, blanc sur fond rouge (toutes les ILV)
    if ean:
        _draw_fitted(page, font, f"EAN : {ean}", Z["ean"], _COL_WHITE,
                     max_size=17, align="left")

    # Prix produit (hors services) — sous le code EAN, blanc sur fond rouge
    if prix > 0:
        prix_str = fmt_ml(prix) + " €"  # ex: "259,98 €"
        _draw_fitted(page, font, prix_str, Z["prix_produit"], _COL_WHITE,
                     max_size=17, align="left")

    # Mensualité ENTIER — grosse, blanche, fond rouge (auto-size pour remplir)
    _draw_fitted(page, font, mensu_entier, Z["mensu_int"], _COL_WHITE,
                 max_size=230, align="right", fill_h=1.0)

    # Centimes — taille fixe (agrandie un peu), blanche, centrée dans sa boîte
    _draw_fitted(page, font, centimes, Z["centimes"], _COL_WHITE,
                 max_size=58, align="center")

    # TAEG — taille fixe à droite du label "TAEG fixe :"
    s = F["taeg"]
    _draw_at(page, font, taeg, s["x"], s["baseline"], s["size"], _COL_WHITE, s["max_w"])

    # Montant total dû — taille fixe à droite du label
    s = F["montant"]
    _draw_at(page, font, total_du, s["x"], s["baseline"], s["size"], _COL_WHITE, s["max_w"])

    # Prix services selon variant (rouge BUT)
    if variant == "gar_liv":
        s = F["liv"]
        _draw_at(page, font, liv_prix, s["x"], s["baseline"], s["size"], _COL_RED, s["max_w"])
        s = F["gar_gl"]
        _draw_at(page, font, gar_prix, s["x"], s["baseline"], s["size"], _COL_RED, s["max_w"])
    elif variant == "gar":
        s = F["gar_g"]
        _draw_at(page, font, gar_prix, s["x"], s["baseline"], s["size"], _COL_RED, s["max_w"])

    # Badge IR (10x, 20x) — le nombre en ROUGE dans le cercle blanc, le « € » est statique.
    # On affiche les décimales quand elles existent (précision : ex. 244,42 et non 244).
    if slug in ("10x", "20x") and interets and interets > 0:
        cents = int(round((interets - int(interets)) * 100))
        badge_num = f"{int(interets)}" if cents == 0 else f"{int(interets)},{cents:02d}"
        # Cercle blanc : nombre rouge, aligné à droite juste avant le « € » statique (x≈732)
        _draw_fitted(page, font, badge_num, Z["badge_ir"], _COL_RED,
                     max_size=58, align="right", fill_h=1.0)

    # Mentions légales — boîte du bas, AU-DESSUS des logos BUT/Cetelem (y=1111).
    # On cherche la plus grande taille qui tient (≤9) sans déborder sur les logos.
    if ml_text:
        ml_zone = fitz.Rect(30, 982, 812, 1105)
        # Conversion Unicode → WinAnsi (CP1252) pour Helvetica built-in
        ml_win = ml_text.replace("€", chr(0x80))    # € → CP1252 0x80
        ml_win = ml_win.replace("–", chr(0x96)) # – (en-dash) → CP1252 0x96
        ml_win = ml_win.replace(" ", " ")       # espace fine insécable → espace normale
        for size in (9, 8.5, 8, 7.5, 7, 6.5, 6, 5.5, 5, 4.5):
            page.draw_rect(ml_zone, color=(1, 1, 1), fill=(1, 1, 1))  # effacer l'essai précédent
            rc = page.insert_textbox(ml_zone, ml_win, fontsize=size, fontname="helv",
                                     color=_COL_DARK, align=fitz.TEXT_ALIGN_JUSTIFY,
                                     encoding=fitz.TEXT_ENCODING_LATIN)
            if rc >= 0:
                break   # taille OK et déjà dessinée

    # Sauvegarder
    os.makedirs(os.path.dirname(str(output_path)), exist_ok=True)
    doc.save(str(output_path), garbage=3, deflate=True)
    doc.close()


# ── Génération PDF A3 — templates STC (100 % vectorisés, sans texte extractible) ──
#
# Ces templates n'ont aucun placeholder textuel scannable.
# Stratégie : couvrir chaque zone placeholder avec un rect de la bonne couleur,
# puis insérer la valeur en texte — et écrire les ML complètes dans la zone blanche.

# Coordonnées fixes pour BDX_2610-...-5X-A3-STC_001.pdf
# (déduites par analyse pixel des paths vectoriels)
_STC_5X_SPECS = {
    # (rect_à_couvrir, origin_insert, font_res, fontsize, color)
    # color = (R,G,B) normalisé 0-1
    # La couleur de fond rouge du template
    "mensualite_entier": dict(
        cover=fitz.Rect(55, 420, 415, 578), origin=(60, 568),
        font="TT1", sz_base=203.4,
        fg=(1, 1, 1), bg=(226/255, 0, 10/255),   # blanc sur rouge
        auto_scale=True,                           # réduire si > 2 chiffres
    ),
    "centimes_str": dict(
        cover=fitz.Rect(418, 498, 530, 578), origin=(422, 568),
        font="TT1", sz_base=95.6,
        fg=(1, 1, 1), bg=(226/255, 0, 10/255),
    ),
    "taeg_str": dict(
        cover=fitz.Rect(72, 592, 256, 618), origin=(75, 613),
        font="TT2", sz_base=22.4,
        fg=(1, 1, 1), bg=(226/255, 0, 10/255),
        prefix="TAEG fixe : ",
    ),
    "total_du_str": dict(
        cover=fitz.Rect(72, 632, 335, 658), origin=(75, 653),
        font="TT2", sz_base=22.4,
        fg=(1, 1, 1), bg=(226/255, 0, 10/255),
        prefix="Montant total dû : ",
    ),
    # Désignation article (bulle saumon, y≈277-300)
    "desig": dict(
        cover=fitz.Rect(92.5, 276.6, 446.7, 300.2), origin=(95, 296.5),
        font="TT1", sz_base=18.0,
        fg=(64/255, 64/255, 66/255), bg=(252/255, 209/255, 192/255),  # gris sur saumon
    ),
    # Prix livraison (y≈785-804, "00€")
    "liv_prix": dict(
        cover=fitz.Rect(242.9, 785, 301.7, 804), origin=(245, 802),
        font="TT1", sz_base=14.0,
        fg=(64/255, 64/255, 66/255), bg=(252/255, 209/255, 192/255),  # gris sur saumon
    ),
    # Prix garantie (y≈786-803, "000€")
    "gar_prix": dict(
        cover=fitz.Rect(569.5, 786, 768.5, 803), origin=(572, 801),
        font="TT1", sz_base=14.0,
        fg=(64/255, 64/255, 66/255), bg=(252/255, 209/255, 192/255),  # gris sur saumon
    ),
    # Zone ML : y=978-1108, x=18-824 (blanc, écriture complète)
    "_ml_rect": fitz.Rect(18, 978, 824, 1108),
    "_ml_fontsize": 8.0,
}


def generate_a3_stc_pdf(slug, variant, designation, calc, montant_finance,
                         garantie_prix, livraison_prix, duree, famille,
                         template_path, output_path,
                         specs=None):
    """
    Génère un PDF A3 ILV à partir d'un template STC (100 % vectorisé).
    Les valeurs sont peintes par-dessus les zones fixes, et les ML complètes
    sont écrites from scratch dans la zone blanche en bas.
    """
    import tempfile as _tmp

    if specs is None:
        specs = _STC_5X_SPECS

    # Calculer les valeurs
    vals = _a3_compute_values(slug, variant, designation, calc, montant_finance,
                               garantie_prix, livraison_prix, duree, famille)

    # Extraire les polices depuis le template de référence
    ref_path = A3_TEMPLATES_DIR / A3_TEMPLATE_FILES.get(slug, A3_TEMPLATE_FILES["5x"])
    font_data = {}
    if ref_path.exists():
        ref = fitz.open(str(ref_path))
        page_idx = A3_VARIANT_TO_PAGE.get(variant, 1)
        for finfo in ref[min(page_idx, ref.page_count - 1)].get_fonts():
            ext = ref.extract_font(finfo[0])
            if ext[3]:
                font_data[finfo[4]] = ext[3]
        ref.close()

    font_files = {}
    tmp_paths  = []
    for res, data in font_data.items():
        fd, fpath = _tmp.mkstemp(suffix=".ttf")
        os.close(fd)
        with open(fpath, "wb") as f:
            f.write(data)
        font_files[res] = fpath
        tmp_paths.append(fpath)

    try:
        doc  = fitz.open(str(template_path))
        page = doc[0]

        for key, spec in specs.items():
            if key.startswith("_"):
                continue

            val = vals.get(key, "") or ""
            if not val:
                continue

            # Adapter la taille police si auto_scale (grands nombres)
            sz = spec["sz_base"]
            if spec.get("auto_scale"):
                n = len(val.replace(chr(0x80), "").replace("€", "").replace(",", "").strip())
                sz = sz if n <= 2 else (160.0 if n == 3 else 125.0)

            # Couvrir l'ancienne valeur avec le fond
            page.draw_rect(spec["cover"], color=None, fill=spec["bg"])

            # Insérer la nouvelle valeur
            # Les polices embarquées (TT1/TT2) utilisent WinAnsiEncoding → chr(0x80)=€
            # Helv (fallback) utilise Unicode → on convertit
            ff = font_files.get(spec["font"])
            if ff:
                text = spec.get("prefix", "") + val   # garder chr(0x80)
            else:
                text = spec.get("prefix", "") + _fmt_eur(val)  # convertir en €

            kw = dict(fontsize=sz, color=spec["fg"])
            if ff:
                kw["fontfile"] = ff
            else:
                kw["fontname"] = "helv"
            page.insert_text(fitz.Point(*spec["origin"]), text, **kw)

        # ── ML complètes dans la zone blanche ─────────────────────────────
        ml_rect = specs.get("_ml_rect")
        ml_fsz  = specs.get("_ml_fontsize", 8.0)
        if ml_rect is not None:
            ml_text = _build_ml_text(slug, vals)
            rc = page.insert_textbox(
                ml_rect, ml_text,
                fontname="helv", fontsize=ml_fsz,
                color=(0, 0, 0), align=0,
            )
            if rc < 0:
                # Si débordement, réduire légèrement la taille
                page.insert_textbox(
                    ml_rect, ml_text,
                    fontname="helv", fontsize=max(6.5, ml_fsz - 1.0),
                    color=(0, 0, 0), align=0,
                )

        # Sauvegarder
        os.makedirs(os.path.dirname(str(output_path)), exist_ok=True)
        fd, tmp_path = _tmp.mkstemp(suffix=".pdf")
        os.close(fd)
        try:
            doc.save(tmp_path, incremental=False, garbage=3, deflate=True)
            doc.close()
            os.replace(tmp_path, str(output_path))
        except Exception:
            doc.close()
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)
            raise

    finally:
        for fpath in tmp_paths:
            try:
                os.unlink(fpath)
            except OSError:
                pass


# ── Répertoire des templates STC ──────────────────────────────────────────────
# Clé : slug, Valeur : (fichier_template, specs_dict)
A3_STC_TEMPLATE_FILES = {
    "5x": ("BDX_2610-DIVERS-ILV-CREDIT-5X-A3-STC_001.pdf", _STC_5X_SPECS),
}


# ── Programme principal ─────────────────────────────────────────────────────

def main():
    # Arguments CLI
    args = sys.argv[1:]
    xlsx_path = None
    prix_min = 300.0

    i = 0
    while i < len(args):
        if args[i] == "--prix-min" and i + 1 < len(args):
            prix_min = float(args[i + 1])
            i += 2
        elif not args[i].startswith("--"):
            xlsx_path = Path(args[i])
            i += 1
        else:
            i += 1

    if xlsx_path is None:
        xlsx_path = BASE_DIR / "NICOLAS_TARDY_Prix_dépliant (3).xlsx"

    if not xlsx_path.exists():
        print(f"ERREUR : fichier introuvable : {xlsx_path}")
        sys.exit(1)

    if not CHAMPS_JSON.exists():
        print(f"ERREUR : champs_all.json introuvable dans {BASE_DIR}")
        sys.exit(1)

    with open(CHAMPS_JSON) as f:
        champs_all = json.load(f)

    # Dossier de sortie
    out_dir = BASE_DIR / f"ILV_depliant_{date.today():%Y%m%d}"
    out_dir.mkdir(exist_ok=True)
    print(f"Dossier de sortie : {out_dir}")

    # Parse Excel
    print(f"Lecture de {xlsx_path.name} (prix minimum : {prix_min:.0f}€)...")
    products = parse_depliant(xlsx_path, prix_min)
    print(f"  → {len(products)} produits éligibles")

    ok = 0
    ok_a3 = 0
    skip_no_credit = 0
    skip_mensualite = 0
    skip_no_template = 0
    errors = []

    for p in products:
        credit_key = determine_credit(p["prix"], p["univers_ilv"])
        if not credit_key:
            skip_no_credit += 1
            continue

        ct = CREDIT_TYPES[credit_key]
        duree = ct["duree"]
        famille = ct["famille"]
        slug = CREDITKEY_TO_SLUG[credit_key]

        entry = champs_all.get(slug)
        if not entry:
            skip_no_template += 1
            continue

        pdf_template = BASE_DIR / entry["pdf"]
        if not pdf_template.exists():
            skip_no_template += 1
            continue

        war = lookup_warranty_all(p["rayon_ilv"], p["prix"])
        _, garantie_prix = best_garantie(war)

        # Variante : gar si garantie dispo, sinon sans (pas de livraison par défaut)
        if garantie_prix > 0:
            variant = "gar"
        else:
            variant = "sans"

        # Vérification que cette variante existe pour ce type de crédit
        if variant not in ct["variants"]:
            variant = "sans"

        page_num = VARIANT_TO_PAGE[variant]

        # Montant financé
        montant_finance = p["prix"] + (garantie_prix if variant in ("gar", "gar_liv") else 0)
        calc = calculer_credit(duree, famille, montant_finance)

        # Règle légale : mensualité >= 16€
        if calc["mensualite"] < 16:
            skip_mensualite += 1
            continue

        product_data = {
            "designation": p["designation"],
            "ean": p["ean"],
            "prix": p["prix"],
            "garantie": garantie_prix if variant in ("gar", "gar_liv") else 0,
            "livraison": 0,  # pas de livraison en génération batch par défaut
        }

        # Champs de cette page
        field_name_to_role = {
            f["name"]: f["role"]
            for f in entry["data"]["fields"]
            if f["page"] == page_num
        }

        role_values = build_field_values(product_data, calc, duree, famille, variant, page_num)

        filename = make_filename(p["rayon_ilv"], p["ean"], credit_key, variant)
        output_path = out_dir / filename

        try:
            generate_ilv_pdf(slug, page_num, field_name_to_role, role_values,
                             pdf_template, output_path)
            ml_text = generer_ml_text(slug, duree, famille, montant_finance, calc)
            ajouter_ml_overlay(str(output_path), slug, ml_text)
            ok += 1
            print(f"  [{ok:3d}] {filename}  ({p['prix']:.2f}€ → {calc['mensualite']:.2f}€/mois)")
        except Exception as e:
            errors.append((filename, str(e)))

        # ── Génération A3 (si template disponible) ──────────────────────────
        a3_tpl_file = A3_TEMPLATE_FILES.get(slug)
        if a3_tpl_file:
            a3_tpl_path = A3_TEMPLATES_DIR / a3_tpl_file
            if a3_tpl_path.exists():
                a3_filename = make_filename(p["rayon_ilv"], p["ean"],
                                            credit_key + "_a3", variant)
                a3_output = out_dir / a3_filename
                try:
                    generate_a3_pdf(
                        slug, variant, p["designation"], calc,
                        montant_finance, garantie_prix,
                        product_data["livraison"],
                        duree, famille,
                        a3_tpl_path, a3_output,
                        ean=p.get("ean", ""), prix=p.get("prix", 0),
                    )
                    ok_a3 += 1
                except Exception as e:
                    errors.append((a3_filename, f"A3: {e}"))

    print()
    print(f"═══ Résumé ═══════════════════════════════════════════")
    print(f"  ILV A5 générées  : {ok}")
    print(f"  ILV A3 générées  : {ok_a3}")
    print(f"  Sans type crédit : {skip_no_credit}")
    print(f"  Mensualité < 16€ : {skip_mensualite}")
    print(f"  Template manquant: {skip_no_template}")
    if errors:
        print(f"  Erreurs          : {len(errors)}")
        for fname, err in errors[:5]:
            print(f"    • {fname}: {err}")
    print(f"  Dossier          : {out_dir}")


if __name__ == "__main__":
    main()
