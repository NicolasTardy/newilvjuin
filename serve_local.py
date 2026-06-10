"""Serveur local combiné : static (racine projet) + API Flask."""
import importlib.util
import io
import json
import os
import tempfile
import zipfile
from datetime import datetime
from pathlib import Path

from flask import Flask, abort, jsonify, request, send_file, send_from_directory, redirect

ROOT = Path(__file__).resolve().parent
DATA_DIR = ROOT / "data"
BACKUPS_DIR = DATA_DIR / "backups"
PRODUCTS_PATH = DATA_DIR / "products.json"
DEPLIANT_PATH = DATA_DIR / "depliant.json"
CHAMPS_JSON = ROOT / "champs_all.json"

ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "")
PORT = int(os.environ.get("PORT", "8766"))
BACKUPS_KEEP = 10

app = Flask(__name__, static_folder=None)

# ── Import du module de génération ────────────────────────────────────────────

def _load_gen():
    spec = importlib.util.spec_from_file_location(
        "gen", ROOT / "generate_ilv_depliant.py"
    )
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod

try:
    gen = _load_gen()
    _champs_all = json.loads(CHAMPS_JSON.read_text(encoding="utf-8"))
    print(f"  Module génération chargé — {len(_champs_all)} slugs")
except Exception as e:
    gen = None
    _champs_all = {}
    print(f"  ATTENTION : module génération non chargé — {e}")


# ── Static files ──────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return redirect("/webapp-ilv.html")

@app.route("/<path:filename>")
def static_files(filename):
    resp = send_from_directory(ROOT, filename)
    # Désactive le cache pour que les modifs HTML/JS soient toujours fraîches
    resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate"
    resp.headers["Pragma"] = "no-cache"
    return resp


# ── Helpers admin ─────────────────────────────────────────────────────────────

def _check_auth():
    pwd = request.headers.get("X-Admin-Password", "")
    if ADMIN_PASSWORD and pwd != ADMIN_PASSWORD:
        abort(401, description="Invalid admin password")


def _atomic_write_json(path: Path, data):
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    with tmp.open("w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    tmp.replace(path)


def _rotate_backup(src: Path, prefix: str):
    if not src.exists():
        return
    BACKUPS_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    (BACKUPS_DIR / f"{prefix}_{ts}.json").write_bytes(src.read_bytes())
    for old in sorted(BACKUPS_DIR.glob(f"{prefix}_*.json"))[:-BACKUPS_KEEP]:
        old.unlink()


# ── API admin ─────────────────────────────────────────────────────────────────

@app.get("/api/health")
def health():
    return {"ok": True, "products_exists": PRODUCTS_PATH.exists(),
            "gen_loaded": gen is not None}


@app.post("/api/upload-products")
def upload_products():
    _check_auth()
    payload = request.get_json(silent=True)
    if not isinstance(payload, list) or len(payload) == 0:
        abort(400, description="Expected non-empty JSON array")
    _rotate_backup(PRODUCTS_PATH, "products")
    _atomic_write_json(PRODUCTS_PATH, payload)
    return {"ok": True, "count": len(payload)}


@app.post("/api/upload-depliant")
def upload_depliant():
    _check_auth()
    payload = request.get_json(silent=True)
    if not isinstance(payload, list):
        abort(400, description="Expected JSON array")
    _rotate_backup(DEPLIANT_PATH, "depliant")
    _atomic_write_json(DEPLIANT_PATH, payload)
    return {"ok": True, "count": len(payload)}


# ── Génération PDF : logique partagée ─────────────────────────────────────────

def _generate_one(item: dict) -> list[tuple[str, bytes]]:
    """
    Génère A5 + A3 pour un produit+crédit donné.
    Retourne une liste de (nom_fichier, bytes_pdf).
    Lève ValueError si paramètres invalides.
    """
    designation    = str(item.get("designation", "")).strip()
    prix           = float(item.get("prix", 0))
    rayon_ilv      = str(item.get("rayon_ilv", "")).strip()
    ean            = str(item.get("ean", "")).strip()
    credit_key     = item.get("credit_key") or None
    include_gar    = bool(item.get("include_garantie", False))
    garantie_prix  = float(item.get("garantie_prix", 0)) if include_gar else 0.0
    include_liv    = bool(item.get("include_livraison", False))
    livraison_prix = float(item.get("livraison_prix", 0)) if include_liv else 0.0

    if not designation or prix <= 0:
        raise ValueError("Désignation et prix requis")

    univers_ilv = gen.RAYON_TO_UNIVERS_ILV.get(rayon_ilv)
    if not univers_ilv:
        # Accepter l'univers passé par le frontend, sinon fallback GEM/TV
        univers_ilv = str(item.get("univers_ilv", "")).strip() or "GEM/TV"

    if not credit_key:
        credit_key = gen.determine_credit(prix, univers_ilv)
    if not credit_key:
        raise ValueError("Aucun crédit applicable")

    # Plafond légal : 5× sans frais limité à 3 000 €
    if credit_key == "5x_sf" and prix > 3000:
        raise ValueError("Le 5× sans frais est limité à 3 000 € maximum")

    ct      = gen.CREDIT_TYPES[credit_key]
    duree   = ct["duree"]
    famille = ct["famille"]
    slug    = gen.CREDITKEY_TO_SLUG[credit_key]

    # Variante : sélectionner le bon template selon les services inclus
    has_gar = include_gar and garantie_prix > 0
    has_liv = include_liv and livraison_prix > 0

    if has_gar and has_liv and "gar_liv" in ct["variants"]:
        # Garantie + Livraison incluses
        variant = "gar_liv"
    elif has_gar and "gar" in ct["variants"]:
        # Garantie seule (pas de livraison, ou livraison ignorée)
        variant = "gar"
        livraison_prix = 0.0  # Ignorer livraison si variante gar
    elif has_liv and "gar_liv" in ct["variants"]:
        # Livraison seule → utiliser gar_liv (mais sans garantie)
        variant = "gar_liv"
        garantie_prix = 0.0  # Pas de garantie
    else:
        # Aucun service sélectionné
        variant = "sans"
        garantie_prix  = 0.0
        livraison_prix = 0.0

    # Montant financé
    montant_finance = prix
    if variant in ("gar", "gar_liv"):
        montant_finance += garantie_prix
    if variant == "gar_liv":
        montant_finance += livraison_prix

    calc = gen.calculer_credit(duree, famille, montant_finance)
    if calc["mensualite"] < 16:
        raise ValueError(f"Mensualité {calc['mensualite']:.2f}€ < 16€ minimum légal")

    results = []
    ean_safe = ean or designation[:12].replace(" ", "_")

    # ── A3 uniquement ─────────────────────────────────────────────────────────
    # Utiliser les templates PDF_MODIFIABLES_CREDIT (variantes -1, -2, -3)
    # Note: Les anciens templates STC sont désactivés en faveur des nouvelles variantes
    a3_tpl_file = gen.A3_TEMPLATE_FILES.get(slug)
    if not a3_tpl_file:
        raise ValueError(f"Pas de template A3 pour le crédit {credit_key} ({slug}x) — template manquant")

    # Ajouter le suffixe variant (-1, -2, -3) au chemin du fichier
    variant_suffix = gen.A3_VARIANT_TO_SUFFIX.get(variant, "-1")
    a3_tpl_path = gen.A3_TEMPLATES_DIR / (a3_tpl_file + variant_suffix + ".pdf")
    if not a3_tpl_path.exists():
        raise ValueError(f"Fichier template introuvable : {a3_tpl_path.name} — variant {variant}")

    fname_a3 = gen.make_filename(rayon_ilv, ean_safe, credit_key + "_a3", variant)
    with tempfile.TemporaryDirectory() as tmpdir:
        out_a3 = Path(tmpdir) / fname_a3
        gen.generate_a3_pdf(
            slug, variant, designation, calc,
            montant_finance, garantie_prix, livraison_prix,
            duree, famille, a3_tpl_path, out_a3, ean=ean, prix=prix,
        )
        results.append((fname_a3, out_a3.read_bytes()))

    return results


# ── API génération unitaire ────────────────────────────────────────────────────

@app.post("/api/generate-ilv")
def generate_ilv():
    if gen is None:
        return jsonify(error="Module de génération non disponible"), 503

    item = request.get_json(silent=True) or {}
    try:
        files = _generate_one(item)
    except ValueError as e:
        return jsonify(error=str(e)), 400
    except Exception as e:
        return jsonify(error=f"Erreur génération : {e}"), 500

    if not files:
        return jsonify(error="Aucun PDF généré — template introuvable"), 500

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for fname, data in files:
            zf.writestr(fname, data)
    buf.seek(0)
    ean = str(item.get("ean", "produit"))
    return send_file(buf, mimetype="application/zip",
                     as_attachment=True,
                     download_name=f"ILV_{ean}.zip")


@app.post("/api/generate-ilv-preview")
def generate_ilv_preview():
    """Génère et retourne le PDF seul pour prévisualisation."""
    if gen is None:
        return jsonify(error="Module de génération non disponible"), 503

    item = request.get_json(silent=True) or {}
    try:
        files = _generate_one(item)
    except ValueError as e:
        return jsonify(error=str(e)), 400
    except Exception as e:
        return jsonify(error=f"Erreur génération : {e}"), 500

    if not files:
        return jsonify(error="Aucun PDF généré"), 500

    # Retourner le premier PDF (seul A3)
    fname, data = files[0]
    return send_file(io.BytesIO(data), mimetype="application/pdf",
                     as_attachment=False,
                     download_name=fname)


# ── API admin : génération auto depuis Excel ──────────────────────────────────

@app.post("/api/admin/generate-from-excel")
def admin_generate_from_excel():
    """Upload an Excel file → auto-generate ILVs with warranty, no delivery."""
    if gen is None:
        return jsonify(error="Module de génération non disponible"), 503
    _check_auth()

    if "file" not in request.files:
        return jsonify(error="Fichier Excel requis"), 400
    f = request.files["file"]

    try:
        import openpyxl
    except ImportError:
        return jsonify(error="openpyxl non installé sur le serveur"), 500

    try:
        wb = openpyxl.load_workbook(f, data_only=True)
    except Exception as e:
        return jsonify(error=f"Impossible de lire le fichier Excel : {e}"), 400

    # Find the data sheet (first non-Synthèse sheet, or active)
    ws = None
    for name in wb.sheetnames:
        if name.lower() != "synthèse":
            ws = wb[name]
            break
    if ws is None:
        ws = wb.active

    # Detect columns by header
    headers = {}
    for col in range(1, ws.max_column + 1):
        h = str(ws.cell(1, col).value or "").strip()
        headers[h] = col

    col_famille = headers.get("LIB_FAMILLE")
    col_ean     = headers.get("EAN_13")
    col_nom     = headers.get("LIB_PRODUIT")
    col_pvd     = headers.get("Prix de Vente à Date du produit TTC")
    col_domaine = headers.get("LIB_DOMAINE")
    col_rayon   = headers.get("LIB_RAYON")

    if not col_nom or not col_pvd:
        return jsonify(error="Colonnes LIB_PRODUIT et Prix de Vente à Date du produit TTC requises"), 400

    # Parse request options
    gar_level = int(request.form.get("gar_level", "2"))  # default 3 étoiles

    import sys
    print(f"[admin-excel] Processing {ws.max_row - 1} rows, gar_level={gar_level}", flush=True)

    buf = io.BytesIO()
    errors = []
    skipped = []
    count = 0

    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for row in range(2, ws.max_row + 1):
            famille_excel = str(ws.cell(row, col_famille).value or "").strip() if col_famille else ""
            ean_raw       = ws.cell(row, col_ean).value if col_ean else None
            designation   = str(ws.cell(row, col_nom).value or "").strip()
            prix_raw      = ws.cell(row, col_pvd).value

            if not designation or not prix_raw:
                continue
            try:
                prix = float(prix_raw)
            except (ValueError, TypeError):
                continue
            if prix <= 0:
                continue

            # EAN normalization
            if ean_raw is not None:
                ean = str(int(ean_raw)) if isinstance(ean_raw, float) else str(ean_raw).strip()
            else:
                ean = ""

            # Map famille Excel → rayonILV
            rayon_ilv = gen.EXCEL_FAMILLE_TO_RAYON_ILV.get(famille_excel, "")
            if not rayon_ilv:
                # Fallback: try rayon Excel
                rayon_excel = str(ws.cell(row, col_rayon).value or "").strip() if col_rayon else ""
                rayon_ilv = gen.EXCEL_FAMILLE_TO_RAYON_ILV.get(rayon_excel, "")
            if not rayon_ilv:
                skipped.append(f"row {row}: famille '{famille_excel}' non mappée")
                continue

            univers_ilv = gen.RAYON_TO_UNIVERS_ILV.get(rayon_ilv, "")
            if not univers_ilv:
                skipped.append(f"row {row}: rayon '{rayon_ilv}' sans univers")
                continue

            # Determine credit type
            credit_key = gen.determine_credit(prix, univers_ilv)
            if not credit_key:
                skipped.append(f"row {row}: pas de crédit pour {prix}€ / {univers_ilv}")
                continue

            # Lookup warranty — incluse par défaut seulement à partir du 20x
            war = gen.lookup_warranty_all(rayon_ilv, prix)
            gar_prix = 0
            gar_default = not credit_key.startswith("5x") and not credit_key.startswith("10x")
            if gar_default:
                if gar_level == 2 and war["gar2"] > 0:
                    gar_prix = war["gar2"]
                elif gar_level == 1 and war["gar1"] > 0:
                    gar_prix = war["gar1"]
                elif gar_level == 2 and war["gar1"] > 0:
                    gar_prix = war["gar1"]

            # Build item for _generate_one (warranty only, no livraison)
            item = {
                "designation":       designation,
                "prix":              prix,
                "rayon_ilv":         rayon_ilv,
                "univers_ilv":       univers_ilv,
                "ean":               ean,
                "credit_key":        credit_key,
                "include_garantie":  gar_prix > 0,
                "garantie_prix":     gar_prix,
                "include_livraison": False,
                "livraison_prix":    0,
            }

            try:
                files = _generate_one(item)
                for fname, data in files:
                    zf.writestr(fname, data)
                    count += 1
            except Exception as e:
                errors.append(f"{ean or designation[:20]}: {e}")

    print(f"[admin-excel] Done: {count} PDFs, {len(errors)} erreurs, {len(skipped)} skipped", flush=True)
    if skipped:
        for s in skipped[:10]:
            print(f"  SKIP: {s}", flush=True)

    if count == 0:
        msg = "Aucun PDF généré"
        if errors:
            msg += " — " + "; ".join(errors[:5])
        if skipped:
            msg += f" ({len(skipped)} produits non mappés)"
        return jsonify(error=msg), 500

    buf.seek(0)
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    resp = send_file(buf, mimetype="application/zip",
                     as_attachment=True,
                     download_name=f"ILV_Excel_{ts}.zip")
    resp.headers["X-ILV-Count"] = str(count)
    resp.headers["X-ILV-Skipped"] = str(len(skipped))
    resp.headers["X-ILV-Errors"] = json.dumps(errors[:20], ensure_ascii=False) if errors else "[]"
    resp.headers["Access-Control-Expose-Headers"] = "X-ILV-Count, X-ILV-Skipped, X-ILV-Errors"
    return resp


# ── API génération panier (batch) ──────────────────────────────────────────────

@app.post("/api/generate-ilv-batch")
def generate_ilv_batch():
    if gen is None:
        return jsonify(error="Module de génération non disponible"), 503

    items = request.get_json(silent=True)
    if not isinstance(items, list) or not items:
        return jsonify(error="Liste de produits requise"), 400

    buf = io.BytesIO()
    errors = []
    count = 0

    import sys
    print(f"[batch] {len(items)} items reçus", flush=True)
    for it in items:
        print(f"  ITEM: {it}", flush=True)
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for item in items:
            print(f"  → rayon_ilv={item.get('rayon_ilv')!r}  credit_key={item.get('credit_key')!r}  prix={item.get('prix')}  gar={item.get('garantie_prix')}", flush=True)
            try:
                files = _generate_one(item)
                print(f"     OK: {[f for f,_ in files]}", flush=True)
                for fname, data in files:
                    zf.writestr(fname, data)
                    count += 1
            except Exception as e:
                print(f"     ERREUR: {e}", flush=True)
                label = item.get("ean") or item.get("designation", "?")[:20]
                errors.append(f"{label}: {e}")

    if count == 0:
        msg = "Aucun PDF généré"
        if errors:
            msg += " — " + "; ".join(errors[:3])
        return jsonify(error=msg), 500

    buf.seek(0)
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    resp = send_file(buf, mimetype="application/zip",
                     as_attachment=True,
                     download_name=f"ILV_{ts}.zip")
    if errors:
        resp.headers["X-ILV-Errors"] = json.dumps(errors, ensure_ascii=False)
        resp.headers["Access-Control-Expose-Headers"] = "X-ILV-Errors"
    resp.headers["X-ILV-Count"] = str(count)
    resp.headers["Access-Control-Expose-Headers"] = "X-ILV-Errors, X-ILV-Count"
    return resp


if __name__ == "__main__":
    print(f"\n  ILV Crédit BUT — serveur local")
    print(f"  → http://localhost:{PORT}/webapp-ilv.html")
    print(f"  → Ctrl+C pour arrêter\n")
    app.run(host="127.0.0.1", port=PORT, debug=False)
